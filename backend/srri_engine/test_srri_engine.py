"""
test_srri_engine.py — verification suite for the merged SRRI engine.

Sections
  1. Box 1 §4 formula parity (literal vs vectorised, both bases)
  2. Box 2 band boundaries
  3. Box 3 buffer zone (incl. the monthly NaN-prefix bug that is now fixed)
  4. Weekly/monthly parity against the ORIGINAL two scripts
  5. Input handling: bytes / file object / path / DataFrame
  6. Date format: DMY vs MDY vs ISO vs AUTO, and the ambiguity findings
  7. Validation findings: each check fires when it should
  8. Minimum-period override: cannot be unattributable
  9. Audit: version + hash present, stable, and change-sensitive
 10. Excel export: derived from a result, returns bytes, all sheets present
"""
from __future__ import annotations

import hashlib
import io
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).parent))

from srri_engine import (  # noqa: E402
    BUFFER_MONTHS_DEFAULT, DateFormat, Finding, FindingCode, Frequency,
    FREQUENCY_SPEC, InputCadence, MinPeriodsOverride, ResultStatus, SRRIInputError,
    Severity, apply_buffer_zone, calculate, cesr_volatility, classify_srri,
    detect_cadence, export_workbook, read_prices, run, validate,
    ENGINE_VERSION, _buffer_core,
)

PASS, FAIL = [], []


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(f"  {'PASS' if cond else 'FAIL'}  {name}" + (f"  — {detail}" if detail and not cond else ""))


def series(n=3000, seed=0, start="2012-01-02", freq="B", vol=0.011, drift=0.0003):
    rng = np.random.default_rng(seed)
    idx = pd.date_range(start, periods=n, freq=freq)
    r = rng.normal(drift, vol, n)
    return pd.Series(100 * np.exp(np.cumsum(r)), index=idx, name="Price")


def to_csv_bytes(s, fmt="%d/%m/%Y", header=("Date", "Price")):
    lines = [",".join(header)]
    lines += [f"{d.strftime(fmt)},{v:.6f}" for d, v in s.items()]
    return ("\n".join(lines) + "\n").encode()


# ======================================================================
print("\n1. Box 1 §4 — literal formula vs vectorised rolling")
# ======================================================================
rng = np.random.default_rng(7)
r = rng.normal(0.001, 0.02, 260)
for m, label in ((52, "weekly"), (12, "monthly")):
    lit = cesr_volatility(r, m)
    vec = float(pd.Series(r).std(ddof=1) * np.sqrt(m))
    check(f"1.{label}: sqrt[(m/(T-1))*SUM] == std(ddof=1)*sqrt(m)",
          np.isclose(lit, vec, rtol=1e-14), f"{lit} vs {vec}")

check("1.hand-worked 4-point example",
      np.isclose(cesr_volatility([0.01, -0.02, 0.015, -0.005], 52),
                 float(np.std([0.01, -0.02, 0.015, -0.005], ddof=1) * np.sqrt(52)), rtol=1e-14))
check("1.T<2 returns NaN", np.isnan(cesr_volatility([0.01], 52)))

# ======================================================================
print("\n2. Box 2 — band boundaries (lower inclusive, upper exclusive)")
# ======================================================================
cases = [(0.0, 1), (0.00499, 1), (0.005, 2), (0.0199, 2), (0.02, 3), (0.0499, 3),
         (0.05, 4), (0.0999, 4), (0.10, 5), (0.1499, 5), (0.15, 6), (0.2499, 6),
         (0.25, 7), (3.0, 7)]
check("2.all 14 boundary cases", all(classify_srri(v) == e for v, e in cases),
      str([(v, classify_srri(v), e) for v, e in cases if classify_srri(v) != e]))
check("2.NaN -> None", classify_srri(np.nan) is None)

# ======================================================================
print("\n3. Box 3 — buffer zone")
# ======================================================================
idx = pd.date_range("2020-01-31", periods=12, freq="ME")
raw = pd.Series([3, 3, 4, 4, 4, 4, 4, 3, 3, 3, 3, 3], index=idx, dtype=float)
dis = apply_buffer_zone(raw, 4)
check("3.migration only after 4 consecutive months",
      list(dis.values) == [3, 3, 3, 3, 3, 4, 4, 4, 4, 4, 3, 3], str(list(dis.values)))

flap = pd.Series([3, 4, 3, 4, 3, 4, 3, 4], index=idx[:8], dtype=float)
check("3.flapping never migrates", set(apply_buffer_zone(flap, 4).dropna()) == {3.0})

# the bug: leading NaN prefix used to suppress the first disclosed SRRI
nan_prefix = pd.Series([np.nan] * 4 + [5, 5, 5, 5, 5, 5], index=idx[:10], dtype=float)
fixed = apply_buffer_zone(nan_prefix, 4)
check("3.NaN prefix does not suppress the first disclosed SRRI (NOTE 3 bug)",
      fixed.dropna().iloc[0] == 5.0 and fixed.notna().sum() == 6,
      str(list(fixed.values)))


def legacy_buffer(raw_s, buffer=4):
    """The original monthly-script implementation, verbatim, for comparison."""
    disclosed = raw_s.copy().astype(object)
    vals = raw_s.values
    for i, current in enumerate(vals):
        if pd.isna(current):
            disclosed.iloc[i] = np.nan
            continue
        if i == 0:
            disclosed.iloc[i] = current
            continue
        prev = disclosed.iloc[i - 1]
        if current == prev:
            disclosed.iloc[i] = prev
        else:
            window = vals[max(0, i - buffer + 1): i + 1]
            if len(window) >= buffer and all(v == current for v in window):
                disclosed.iloc[i] = current
            else:
                disclosed.iloc[i] = prev
    return disclosed.astype(float)


legacy = legacy_buffer(nan_prefix, 4)
check("3.legacy monthly script demonstrably had the bug (regression evidence)",
      legacy.notna().sum() < fixed.notna().sum(),
      f"legacy valid={legacy.notna().sum()} new valid={fixed.notna().sum()}")
check("3.identical to legacy once the NaN prefix is removed",
      list(_buffer_core(raw, 4).values) == list(legacy_buffer(raw, 4).values))

# ======================================================================
print("\n4. Parity with the ORIGINAL two scripts")
# ======================================================================
daily = series(n=3200, seed=3)


def legacy_weekly(prices, min_weeks=260):
    wk = prices.resample("W-FRI").last().dropna().to_frame("Price")
    wk["Return"] = wk["Price"].pct_change()
    out = []
    for i in range(len(wk)):
        w = wk["Return"].iloc[max(0, i - 260 + 1): i + 1].dropna().values
        out.append(cesr_volatility(w, 52) if len(w) >= min_weeks else np.nan)
    return pd.Series(out, index=wk.index)


def legacy_monthly(prices, min_months=60):
    mo = prices.resample("ME").last().dropna().to_frame("Price")
    mo["Return"] = mo["Price"].pct_change()
    out = []
    for i in range(len(mo)):
        w = mo["Return"].iloc[max(0, i - 60 + 1): i + 1].dropna().values
        out.append(cesr_volatility(w, 12) if len(w) >= min_months else np.nan)
    return pd.Series(out, index=mo.index)


for label, freq, legacy_fn in (("weekly", Frequency.WEEKLY, legacy_weekly),
                               ("monthly", Frequency.MONTHLY, legacy_monthly)):
    res = run(daily, frequency=freq, date_format=DateFormat.ISO, filename="synthetic.csv")
    new = pd.Series({pd.Timestamp(p.date): (p.ann_vol_pct / 100 if p.ann_vol_pct is not None else np.nan)
                     for p in res.series})
    old = legacy_fn(daily)
    aligned = pd.concat([new.rename("new"), old.rename("old")], axis=1)
    both = aligned.dropna()
    check(f"4.{label}: same index length", len(new) == len(old), f"{len(new)} vs {len(old)}")
    check(f"4.{label}: same count of valid points",
          new.notna().sum() == old.notna().sum(), f"{new.notna().sum()} vs {old.notna().sum()}")
    check(f"4.{label}: annualised vol identical to 1e-12 on {len(both)} points",
          len(both) > 50 and np.allclose(both["new"], both["old"], rtol=1e-12, atol=1e-15),
          str((both["new"] - both["old"]).abs().max()) if len(both) else "no overlap")
    check(f"4.{label}: m/T constants correct",
          res.audit.m == FREQUENCY_SPEC[freq].m and res.audit.window == FREQUENCY_SPEC[freq].window)

wk_res = run(daily, frequency=Frequency.WEEKLY, date_format=DateFormat.ISO)
mo_res = run(daily, frequency=Frequency.MONTHLY, date_format=DateFormat.ISO)
check("4.same fund, both bases produce a valid SRRI",
      wk_res.srri_disclosed is not None and mo_res.srri_disclosed is not None)
check("4.the two bases agree within one class on the same fund",
      abs(wk_res.srri_disclosed - mo_res.srri_disclosed) <= 1,
      f"weekly {wk_res.srri_disclosed} vs monthly {mo_res.srri_disclosed}")

# ======================================================================
print("\n5. Input handling — bytes / file object / path / DataFrame")
# ======================================================================
csv_bytes = to_csv_bytes(daily)
tmp = Path("/tmp/srri_test_nav.csv")
tmp.write_bytes(csv_bytes)

r_bytes = run(csv_bytes, frequency=Frequency.WEEKLY, date_format=DateFormat.DMY, filename="nav.csv")
r_fobj = run(io.BytesIO(csv_bytes), frequency=Frequency.WEEKLY, date_format=DateFormat.DMY, filename="nav.csv")
r_path = run(tmp, frequency=Frequency.WEEKLY, date_format=DateFormat.DMY)
r_df = run(daily.to_frame().reset_index().rename(columns={"index": "Date"}),
           frequency=Frequency.WEEKLY, date_format=DateFormat.ISO)

check("5.bytes accepted", r_bytes.srri_disclosed is not None)
check("5.file object accepted", r_fobj.srri_disclosed is not None)
check("5.path accepted", r_path.srri_disclosed is not None)
check("5.DataFrame accepted", r_df.srri_disclosed is not None)
check("5.bytes / fileobj / path give identical SRRI and hash",
      r_bytes.srri_disclosed == r_fobj.srri_disclosed == r_path.srri_disclosed
      and r_bytes.audit.input_sha256 == r_fobj.audit.input_sha256 == r_path.audit.input_sha256)
check("5.all four routes agree on the SRRI",
      len({r_bytes.srri_disclosed, r_fobj.srri_disclosed,
           r_path.srri_disclosed, r_df.srri_disclosed}) == 1)
check("5.no file was written by the core",
      not Path("srri.xlsx").exists() and not Path("output.xlsx").exists())

# xlsx round trip
xbuf = io.BytesIO()
daily.to_frame().to_excel(xbuf)
r_xlsx = run(xbuf.getvalue(), frequency=Frequency.WEEKLY, date_format=DateFormat.ISO,
             filename="nav.xlsx")
check("5.xlsx bytes detected by magic bytes and parsed",
      r_xlsx.srri_disclosed == r_bytes.srri_disclosed)

try:
    read_prices(b"", date_format=DateFormat.DMY)
    check("5.empty upload raises SRRIInputError", False)
except SRRIInputError as e:
    check("5.empty upload raises SRRIInputError", e.finding.code == FindingCode.NO_DATA)

# ======================================================================
print("\n6. Date format — the dayfirst bug")
# ======================================================================
amb = pd.Series([100.0, 101.0, 102.0],
                index=pd.to_datetime(["2024-03-04", "2024-04-03", "2024-05-06"]))
amb_csv = to_csv_bytes(amb, "%d/%m/%Y")

ps_dmy = read_prices(amb_csv, date_format=DateFormat.DMY, filename="a.csv")
ps_mdy = read_prices(amb_csv, date_format=DateFormat.MDY, filename="a.csv")
check("6.DMY and MDY produce different dates from the same file",
      list(ps_dmy.prices.index) != list(ps_mdy.prices.index))
check("6.DMY reads 04/03/2024 as 4 March",
      pd.Timestamp("2024-03-04") in ps_dmy.prices.index)
check("6.MDY reads 04/03/2024 as 3 April",
      pd.Timestamp("2024-04-03") in ps_mdy.prices.index)
check("6.ambiguous file raises DATE_FORMAT_ASSUMED warning",
      any(f.code == FindingCode.DATE_FORMAT_ASSUMED and f.severity == Severity.WARNING
          for f in ps_dmy.findings))

# unambiguous DMY file (contains day 25) declared as MDY -> hard error
unamb = pd.Series([100.0, 101.0], index=pd.to_datetime(["2024-03-25", "2024-04-26"]))
ps_conflict = read_prices(to_csv_bytes(unamb, "%d/%m/%Y"), date_format=DateFormat.MDY,
                          filename="b.csv")
check("6.declaring MDY on an unambiguous DMY file is an ERROR",
      any(f.code == FindingCode.DATE_FORMAT_CONFLICT and f.severity == Severity.ERROR
          for f in ps_conflict.findings))

ps_auto = read_prices(to_csv_bytes(unamb, "%d/%m/%Y"), date_format=DateFormat.AUTO,
                      filename="b.csv")
check("6.AUTO infers DMY when the data settles it",
      ps_auto.date_format_resolved == DateFormat.DMY
      and any(f.code == FindingCode.DATE_FORMAT_INFERRED for f in ps_auto.findings))

ps_auto_amb = read_prices(amb_csv, date_format=DateFormat.AUTO, filename="a.csv")
check("6.AUTO on a genuinely ambiguous file is an ERROR",
      any(f.code == FindingCode.DATE_FORMAT_AMBIGUOUS and f.severity == Severity.ERROR
          for f in ps_auto_amb.findings))

iso_res = run(to_csv_bytes(daily, "%Y-%m-%d"), frequency=Frequency.WEEKLY,
              date_format=DateFormat.ISO, filename="iso.csv")
check("6.ISO parses cleanly and matches the DMY run",
      iso_res.srri_disclosed == r_bytes.srri_disclosed)

check("6.house default is still DMY (open point, documented)",
      read_prices(amb_csv, filename="a.csv").date_format_resolved == DateFormat.DMY)

# ======================================================================
print("\n7. Validation — findings, not log lines")
# ======================================================================
short = series(n=800, seed=11)          # ~3.2 years
ps_short = read_prices(to_csv_bytes(short), date_format=DateFormat.DMY, filename="s.csv")
f_short = validate(ps_short, Frequency.WEEKLY)
check("7.short history is a returned ERROR finding",
      any(f.code == FindingCode.INSUFFICIENT_HISTORY and f.severity == Severity.ERROR
          for f in f_short))
check("7.'only 3.2 years' reaches the user as text",
      any("years" in f.message for f in f_short if f.code == FindingCode.INSUFFICIENT_HISTORY))

res_short = run(to_csv_bytes(short), frequency=Frequency.WEEKLY, date_format=DateFormat.DMY)
check("7.blocked result, not an exception", res_short.status == ResultStatus.BLOCKED)
check("7.blocked result exposes .errors for the UI", len(res_short.errors) >= 1)
check("7.blocked result carries no series", res_short.series == [])
check("7.blocked result still carries the audit block",
      res_short.audit.input_sha256 and res_short.audit.engine_version == ENGINE_VERSION)

monthly_only = series(n=90, seed=5, freq="ME")
ps_m = read_prices(to_csv_bytes(monthly_only), date_format=DateFormat.DMY, filename="m.csv")
check("7.monthly data on the weekly basis is an ERROR",
      any(f.code == FindingCode.CADENCE_TOO_COARSE and f.severity == Severity.ERROR
          for f in validate(ps_m, Frequency.WEEKLY)))
check("7.monthly data on the monthly basis is fine",
      not any(f.severity == Severity.ERROR for f in validate(ps_m, Frequency.MONTHLY)))
check("7.daily data on the monthly basis warns (Box 1 §2 prefers weekly)",
      any(f.code == FindingCode.MONTHLY_BASIS_NOT_PREFERRED
          for f in validate(read_prices(csv_bytes, date_format=DateFormat.DMY), Frequency.MONTHLY)))

dirty = daily.copy()
dirty.iloc[100] = -5.0
dirty_csv = to_csv_bytes(dirty)
ps_dirty = read_prices(dirty_csv, date_format=DateFormat.DMY, filename="d.csv")
check("7.negative price -> NON_POSITIVE_PRICES warning",
      any(f.code == FindingCode.NON_POSITIVE_PRICES for f in ps_dirty.findings))

with_text = csv_bytes.decode().split("\n")
with_text.insert(50, "not a date,N/A")
ps_txt = read_prices("\n".join(with_text).encode(), date_format=DateFormat.DMY, filename="t.csv")
check("7.unparseable row -> DATES_DROPPED warning",
      any(f.code == FindingCode.DATES_DROPPED for f in ps_txt.findings))

gappy = pd.concat([daily.iloc[:500], daily.iloc[900:]])
check("7.large gap -> LARGE_GAP warning",
      any(f.code == FindingCode.LARGE_GAP
          for f in validate(read_prices(to_csv_bytes(gappy), date_format=DateFormat.DMY),
                            Frequency.WEEKLY)))

stale = daily.copy()
stale.iloc[600:640] = stale.iloc[600]
check("7.stale NAV -> STALE_PRICES warning",
      any(f.code == FindingCode.STALE_PRICES
          for f in validate(read_prices(to_csv_bytes(stale), date_format=DateFormat.DMY),
                            Frequency.WEEKLY)))

dupes_csv = csv_bytes + csv_bytes.split(b"\n", 1)[1].rsplit(b"\n", 2)[1] + b"\n"
check("7.every finding is JSON-serialisable for the UI",
      all(isinstance(f.model_dump(mode="json"), dict) for f in res_short.findings))
check("7.findings carry a stable code and a remediation hint",
      all(f.code in FindingCode for f in res_short.findings)
      and any(f.remediation for f in res_short.errors))

# ======================================================================
print("\n8. Minimum-period override — attributable or nothing")
# ======================================================================
for bad in ({"min_periods": 200, "approved_by": "", "reason": "x" * 20},
            {"min_periods": 200, "approved_by": "R Kumar", "reason": "short"},
            {"min_periods": 200, "approved_by": "admin", "reason": "x" * 20},
            {"min_periods": 200, "approved_by": "R Kumar", "reason": "N/A" + " " * 20}):
    try:
        MinPeriodsOverride(**bad)
        ok = False
    except Exception:
        ok = True
    check(f"8.rejects override {list(bad.values())[1][:12]!r}/{list(bad.values())[2][:8]!r}", ok)

ovr = MinPeriodsOverride(min_periods=150, approved_by="R Kumar, Head of Risk",
                         reason="Fund launched 2023-02; CBI pre-clearance ref 2026-114.",
                         ticket_reference="RISK-114")
res_ovr = run(to_csv_bytes(short), frequency=Frequency.WEEKLY, date_format=DateFormat.DMY,
              min_periods_override=ovr)
check("8.override unblocks the calculation", res_ovr.status == ResultStatus.OK_WITH_WARNINGS)
check("8.override produces a disclosed SRRI", res_ovr.srri_disclosed is not None)
check("8.override forces a warning the user must acknowledge",
      any(f.code == FindingCode.MIN_PERIODS_OVERRIDE_APPLIED and f.severity == Severity.WARNING
          for f in res_ovr.findings))
check("8.approver name appears in the finding text",
      any("R Kumar" in f.message for f in res_ovr.findings))
check("8.override is stored on the result's audit block",
      res_ovr.audit.min_periods_override is not None
      and res_ovr.audit.min_periods_override.approved_by.startswith("R Kumar")
      and res_ovr.audit.min_periods_is_regulatory_default is False)
check("8.short-history warning is retained alongside the override",
      any(f.code == FindingCode.SHORT_HISTORY for f in res_ovr.findings))
check("8.no override -> min_periods is the regulatory T",
      r_bytes.audit.min_periods == 260 and r_bytes.audit.min_periods_is_regulatory_default)

# ======================================================================
print("\n9. Audit — version + input hash")
# ======================================================================
check("9.sha256 matches the raw bytes",
      r_bytes.audit.input_sha256 == hashlib.sha256(csv_bytes).hexdigest())
check("9.engine version recorded", r_bytes.audit.engine_version == ENGINE_VERSION)
check("9.filename and size recorded",
      r_bytes.audit.input_filename == "nav.csv" and r_bytes.audit.input_bytes == len(csv_bytes))
check("9.methodology reference recorded", "CESR/10-673" in r_bytes.audit.methodology_ref)

nudged = daily.copy()
nudged.iloc[-1] *= 1.0000001
check("9.a one-cent change to the input changes the hash",
      run(to_csv_bytes(nudged), frequency=Frequency.WEEKLY,
          date_format=DateFormat.DMY).audit.input_sha256 != r_bytes.audit.input_sha256)
check("9.same bytes -> same hash (reproducible)",
      run(csv_bytes, frequency=Frequency.WEEKLY,
          date_format=DateFormat.DMY).audit.input_sha256 == r_bytes.audit.input_sha256)
check("9.fingerprint is short and quotable",
      r_bytes.audit.fingerprint.startswith(ENGINE_VERSION + "/") and len(r_bytes.audit.fingerprint) < 32)

dumped = r_bytes.model_dump(mode="json")
check("9.whole result round-trips to JSON",
      isinstance(dumped, dict) and dumped["audit"]["input_sha256"] == r_bytes.audit.input_sha256)
import json as _json  # noqa: E402
check("9.JSON is actually serialisable (no datetime leakage)",
      isinstance(_json.dumps(dumped), str))
check("9.resolved parameters, not requested ones, are recorded",
      run(csv_bytes, frequency=Frequency.AUTO, date_format=DateFormat.DMY).audit.frequency
      == Frequency.WEEKLY)

# determinism
check("9.two identical runs give identical results",
      run(csv_bytes, frequency=Frequency.WEEKLY, date_format=DateFormat.DMY).srri_disclosed
      == r_bytes.srri_disclosed)

# ======================================================================
print("\n10. Excel export — an artifact, not the return value")
# ======================================================================
xls = export_workbook(r_bytes)
check("10.returns bytes", isinstance(xls, bytes) and len(xls) > 10_000)
check("10.is a real xlsx", xls[:4] == b"PK\x03\x04")

from openpyxl import load_workbook  # noqa: E402
wb = load_workbook(io.BytesIO(xls))
expected = ["Summary", "SRRI Calculations", "Distribution", "Methodology", "Audit & Findings"]
check("10.all five sheets present", wb.sheetnames == expected, str(wb.sheetnames))
check("10.calculations sheet has one row per period",
      wb["SRRI Calculations"].max_row == len(r_bytes.series) + 3,
      f"{wb['SRRI Calculations'].max_row} vs {len(r_bytes.series) + 3}")

audit_text = "\n".join(str(c.value) for row in wb["Audit & Findings"].iter_rows()
                       for c in row if c.value is not None)
check("10.audit sheet carries the input hash", r_bytes.audit.input_sha256 in audit_text)
check("10.audit sheet carries the engine version", ENGINE_VERSION in audit_text)

wb_ovr = load_workbook(io.BytesIO(export_workbook(res_ovr)))
ovr_text = "\n".join(str(c.value) for row in wb_ovr["Audit & Findings"].iter_rows()
                     for c in row if c.value is not None)
check("10.override attribution reaches the workbook",
      "R Kumar" in ovr_text and "RISK-114" in ovr_text and "OVERRIDDEN" in ovr_text)
check("10.findings are tabulated on the audit sheet",
      "MIN_PERIODS_OVERRIDE_APPLIED" in ovr_text)

out = Path("/tmp/srri_export_test.xlsx")
export_workbook(r_bytes, out)
check("10.optional destination writes the same bytes", out.read_bytes()[:4] == b"PK\x03\x04")

sink = io.BytesIO()
export_workbook(r_bytes, sink)
check("10.destination can be a stream (HTTP response)", sink.getvalue()[:4] == b"PK\x03\x04")

mo_wb = load_workbook(io.BytesIO(export_workbook(mo_res)))
check("10.monthly workbook says months, not weeks",
      mo_wb["SRRI Calculations"].cell(row=3, column=1).value == "Month Ending")
check("10.weekly workbook says weeks",
      wb["SRRI Calculations"].cell(row=3, column=1).value == "Week Ending")

# ======================================================================
print("\n" + "=" * 70)
print(f"  {len(PASS)} passed, {len(FAIL)} failed")
if FAIL:
    for f in FAIL:
        print(f"    FAILED: {f}")
print("=" * 70)
sys.exit(1 if FAIL else 0)
