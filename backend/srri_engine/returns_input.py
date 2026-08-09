"""
returns_input.py — turn a supplied return series into a price series

Some administrators send a percentage return time series rather than NAV
prices.  Both KIID calculations need prices, so the validation step converts
one to the other before either engine sees the file.

    detect_series_kind(values)      -> what the column actually is, with evidence
    to_prices(values, ...)          -> a price series, indexed at 100

Why the base value does not matter
----------------------------------
Every output of both engines is a RATIO of two points on the series — CESR
Box 1 period returns, and the calendar-year returns behind the bar chart.
Multiplying the whole series by a constant changes nothing.  100 is a
convention, not an input to any figure.

Detection, and where it refuses
-------------------------------
The detector reads five properties of the column: whether any value is
negative, how often consecutive values change sign, the median absolute value,
the first value, and the lag-1 autocorrelation of the levels.  Those separate
prices, periodic returns and cumulative returns cleanly in almost every real
file.

What they do NOT always separate is percent from decimal on a very low
volatility series.  A money-market fund quoted in percent (median 0.004 %) and
the same fund quoted as decimals (median 0.00004) sit on the same side of any
fixed threshold you pick.  Where the evidence does not decide, this module
raises `SERIES_KIND_AMBIGUOUS` and refuses, exactly as the date-format handling
does for `03/04/2024`.  A silent guess here scales every figure by 100.

The one convention that has to be stated
----------------------------------------
For a periodic return series, the return on the first row is treated as the
return of the period ENDING on that date.  The 100 base therefore sits one
period before the first date, and the first price is `100 x (1 + r0)`.  The
alternative — dropping the first return so the series opens exactly at 100 —
discards a real observation.  Where the first row is blank, which is common
because the first period has no return, the series simply opens at 100.
"""

from __future__ import annotations

import io
import logging
from enum import Enum
from pathlib import Path
from typing import Any, Optional, Union

import numpy as np
import pandas as pd
from pydantic import BaseModel, ConfigDict, Field

from srri_engine import Finding, FindingCode, Severity

log = logging.getLogger("returns_input")

# --- thresholds, all of them house choices ---------------------------------
SIGN_FLIP_PERIODIC_MIN = 0.15   # periodic returns change sign often
LEVEL_AUTOCORR_PRICE_MIN = 0.90  # a price series is a near-random walk in levels
CUMULATIVE_FIRST_TOL = 1e-9     # a cumulative series starts at zero
CUMULATIVE_DRIFT_RATIO = 5.0    # and wanders far from its typical step

# --------------------------------------------------------------------------
# UNITS DETECTION — percent or decimal?
#
# NOTE, because the numbers look deceptively familiar: these are thresholds on
# the RAW VALUES IN THE UPLOADED FILE, applied before any calculation happens.
# They are NOT volatility thresholds and they have NOTHING to do with the Box 2
# SRRI bands in `srri_engine.SRRI_BANDS` (0 / 0.5 / 2 / 5 / 10 / 15 / 25 %).
# The 0.5 appearing in both is coincidence. Nothing here touches risk
# classification.
#
# The question is only this: a cell reading 1.25 — is that +1.25 % or +125 %?
# Both are conventions administrators use. The largest single move in the
# column decides it.
#
#   >= 1.5      PERCENT.  As decimals that is a +150 % move in one period.
#   0.02 - 0.5  DECIMAL.  As percentages the fund never moved more than 0.5 %
#               in any period across its whole history.
#   0.5 - 1.5   DEAD ZONE. Both readings describe a real fund. Refuse.
#   < 0.02      TOO QUIET. A money-market fund quoted in percent and the same
#               fund quoted in decimals both land here, and the two readings
#               are 100x apart in the published SRRI — 0.06 % annualised is
#               class 1, 6.3 % is class 4. This floor is why the module can
#               refuse at all: without it a quiet percent series resolves
#               silently to decimal.
# --------------------------------------------------------------------------
RAW_VALUE_PERCENT_MIN = 1.5      # at or above this, the column is percentages
RAW_VALUE_DECIMAL_MAX = 0.5      # at or below this, the column is decimals
RAW_VALUE_TOO_QUIET = 0.02       # below this, neither reading can be ruled out

# Backwards-compatible aliases; the names above are the ones to use.
PERCENT_MAX_ABS_MIN = RAW_VALUE_PERCENT_MIN
DECIMAL_MAX_ABS_MAX = RAW_VALUE_DECIMAL_MAX
DECIMAL_MAX_ABS_MIN = RAW_VALUE_TOO_QUIET


class SeriesKind(str, Enum):
    PRICE = "price"
    RETURN_PCT = "return_pct"                   # 1.25 means +1.25 %
    RETURN_DECIMAL = "return_decimal"           # 0.0125 means +1.25 %
    CUMULATIVE_PCT = "cumulative_pct"           # 12.5 means +12.5 % since start
    CUMULATIVE_DECIMAL = "cumulative_decimal"   # 0.125 means +12.5 % since start
    AMBIGUOUS = "ambiguous"


class Detection(BaseModel):
    """What the column looks like, and the evidence for saying so."""
    model_config = ConfigDict(frozen=True)

    kind: SeriesKind
    confident: bool
    reason: str
    candidates: list[SeriesKind] = Field(default_factory=list)
    evidence: dict[str, Any] = Field(default_factory=dict)

    @property
    def is_return_series(self) -> bool:
        return self.kind in (SeriesKind.RETURN_PCT, SeriesKind.RETURN_DECIMAL,
                             SeriesKind.CUMULATIVE_PCT,
                             SeriesKind.CUMULATIVE_DECIMAL)


def _evidence(values: pd.Series) -> dict[str, Any]:
    clean = pd.to_numeric(values, errors="coerce").dropna()
    if clean.empty:
        return {"n": 0}

    nonzero = clean[clean != 0]
    if len(nonzero) > 1:
        signs = np.sign(nonzero.to_numpy())
        flips = float(np.mean(signs[1:] != signs[:-1]))
    else:
        flips = 0.0

    levels = clean.to_numpy(dtype=float)
    if len(levels) > 2 and np.std(levels) > 0:
        autocorr = float(np.corrcoef(levels[:-1], levels[1:])[0, 1])
    else:
        autocorr = float("nan")

    median_abs = float(np.median(np.abs(levels)))
    steps = np.abs(np.diff(levels))
    return {
        "n": int(len(clean)),
        "max_abs_diff": float(np.max(steps)) if len(steps) else 0.0,
        "median_abs_diff": float(np.median(steps)) if len(steps) else 0.0,
        "any_negative": bool((clean < 0).any()),
        "any_zero": bool((clean == 0).any()),
        "min": float(clean.min()),
        "max": float(clean.max()),
        "first": float(clean.iloc[0]),
        "last": float(clean.iloc[-1]),
        "median_abs": median_abs,
        "max_abs": float(np.max(np.abs(levels))),
        "sign_flip_rate": round(flips, 4),
        "level_autocorr": None if np.isnan(autocorr) else round(autocorr, 4),
    }


def _looks_numeric(value: Any) -> bool:
    """True for a number, or for text that is one — including `"1.25%"`."""
    if isinstance(value, bool):
        return False
    if isinstance(value, (int, float)):
        return True
    if isinstance(value, str):
        try:
            float(value.strip().rstrip("%").replace(",", ""))
            return True
        except ValueError:
            return False
    return False


def _to_number(value: Any) -> float:
    """Coerce a cell to a float, tolerating a trailing % and thousands commas.

    The % is stripped, NOT divided out — the scale divisor is decided once, by
    `detect_units_from_workbook`, and applied in one place.
    """
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.strip().rstrip("%").replace(",", ""))
        except ValueError:
            return float("nan")
    return float("nan")


class UnitsHint(BaseModel):
    """What the workbook itself says about the column, before any statistics.

    A declared `%` beats any magnitude heuristic, so this is consulted first.
    """
    model_config = ConfigDict(frozen=True)

    kind: Optional[SeriesKind] = None
    scale_divisor: float = 1.0
    source: str = "none"          # cell_format | value_suffix | header_text
    header: Optional[str] = None
    confident: bool = False
    evidence: dict[str, Any] = Field(default_factory=dict)


# Header words that mean "this column is a return, not a price".
_RETURN_WORDS = ("return", "perf", "chg", "change", "%")
_PRICE_WORDS = ("nav", "price", "close", "value")
_CUMULATIVE_WORDS = ("cumulative", "cumul", "since inception", "itd", "total return")


def detect_units_from_workbook(source: Union[str, bytes, "Path"],
                               *,
                               sheet: Optional[Union[str, int]] = None,
                               max_scan_rows: int = 60) -> UnitsHint:
    """Read what the FILE declares about its own units.

    Three signals, in strict precedence order, because two of them mean
    opposite things about the stored number:

    1. **Cell number format contains `%`** — Excel percent formatting. The
       stored value is ALREADY the decimal: a cell showing `1.25%` holds
       `0.0125`. Divisor 1. This wins because it is a property of the stored
       value, not a label.
    2. **Values are text ending in `%`** — e.g. the string `"1.25%"`. The
       number is in percent units. Divisor 100.
    3. **Header text contains `%` or a return word** — e.g. `Return %`. The
       values are plain numbers in percent units. Divisor 100.

    Reading (1) as if it were (3) divides an already-decimal series by 100
    again, and understates volatility a hundredfold. That is the whole reason
    this function exists rather than a regex on the header.

    Returns a hint with `kind=None` where the file declares nothing, in which
    case the caller falls back to `detect_series_kind()`.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:                                    # pragma: no cover
        return UnitsHint(source="none",
                         evidence={"error": "openpyxl unavailable"})

    try:
        src = io.BytesIO(source) if isinstance(source, bytes) else str(source)
        wb = load_workbook(src, data_only=True)
        ws = wb[sheet] if isinstance(sheet, str) else (
            wb.worksheets[sheet] if isinstance(sheet, int) else wb.worksheets[0])
    except Exception as exc:                               # noqa: BLE001
        return UnitsHint(source="none",
                         evidence={"error": f"{type(exc).__name__}: {exc}"})

    # Find the header row: the first row whose cells are text and which is
    # followed by numbers. The fixed template puts it on row 7.
    header_row = None
    header_text = None
    value_col = None
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        labels = [(c, ws.cell(row=r, column=c).value)
                  for c in range(1, min(ws.max_column, 12) + 1)]
        texts = [(c, str(v)) for c, v in labels if isinstance(v, str) and v.strip()]
        if len(texts) < 2:
            continue
        below = [ws.cell(row=r + 1, column=c).value for c, _ in texts]
        # A return column may hold TEXT like "1.25%", so a numeric cell below
        # cannot be required — that would skip the header row entirely and the
        # file would look like it declared nothing.
        if not any(_looks_numeric(v) for v in below):
            continue
        # The value column is the one that is not the date column.
        for c, text in texts:
            low = text.lower()
            if "date" in low:
                continue
            value_col, header_text, header_row = c, text, r
            break
        if header_row:
            break

    if value_col is None:
        return UnitsHint(source="none", evidence={"reason": "no header found"})

    fmts, str_pct, numeric_seen = [], 0, 0
    for r in range(header_row + 1, min(header_row + 1 + max_scan_rows, ws.max_row) + 1):
        cell = ws.cell(row=r, column=value_col)
        if cell.value is None:
            continue
        if isinstance(cell.value, str) and cell.value.strip().endswith("%"):
            str_pct += 1
        if isinstance(cell.value, (int, float)):
            numeric_seen += 1
            fmts.append(cell.number_format or "")

    pct_fmt = sum(1 for f in fmts if "%" in f)
    low_header = (header_text or "").lower()
    header_says_return = any(w in low_header for w in _RETURN_WORDS)
    header_says_price = any(w in low_header for w in _PRICE_WORDS)
    cumulative = any(w in low_header for w in _CUMULATIVE_WORDS)

    ev = {"header": header_text, "header_row": header_row,
          "value_column": value_col, "numeric_cells_sampled": numeric_seen,
          "percent_formatted_cells": pct_fmt,
          "text_percent_cells": str_pct,
          "header_says_return": header_says_return,
          "header_says_price": header_says_price}

    # 1 — Excel percent formatting. Stored value is already decimal.
    if numeric_seen and pct_fmt >= max(1, int(0.8 * numeric_seen)):
        return UnitsHint(
            kind=(SeriesKind.CUMULATIVE_DECIMAL if cumulative
                  else SeriesKind.RETURN_DECIMAL),
            scale_divisor=1.0, source="cell_format", header=header_text,
            confident=True,
            evidence={**ev, "note": ("Cells are percent-formatted, so the "
                                     "stored value is already the decimal — "
                                     "1.25 % is stored as 0.0125. No division "
                                     "by 100.")})

    # 2 — text values like "1.25%".
    if str_pct >= 3:
        return UnitsHint(
            kind=(SeriesKind.CUMULATIVE_PCT if cumulative
                  else SeriesKind.RETURN_PCT),
            scale_divisor=100.0, source="value_suffix", header=header_text,
            confident=True,
            evidence={**ev, "note": ("Values carry a literal % suffix, so the "
                                     "number is in percent units.")})

    # 3 — the header says so.
    if header_says_return and not header_says_price:
        return UnitsHint(
            kind=(SeriesKind.CUMULATIVE_PCT if cumulative
                  else SeriesKind.RETURN_PCT),
            scale_divisor=100.0, source="header_text", header=header_text,
            confident=True,
            evidence={**ev, "note": (f"The header {header_text!r} names a "
                                     f"return, and the values are plain "
                                     f"numbers — percent units.")})

    if header_says_price:
        return UnitsHint(kind=SeriesKind.PRICE, scale_divisor=1.0,
                         source="header_text", header=header_text,
                         confident=True,
                         evidence={**ev, "note": "The header names a price."})

    return UnitsHint(source="none", header=header_text, evidence=ev)


def _percent_or_decimal(magnitude: float, noun: str
                        ) -> tuple[Optional[str], str]:
    """Decide percent vs decimal from the largest single move, or refuse.

    Returns ("pct" | "dec" | None, explanation).  None means the two readings
    are 100x apart in the published figure and the data cannot separate them.
    """
    if magnitude >= RAW_VALUE_PERCENT_MIN:
        return "pct", (f"The largest {noun} is {magnitude:.3g} — percentages. "
                       f"Read as decimals that is a {magnitude:.0%} move in a "
                       f"single period.")
    if RAW_VALUE_TOO_QUIET <= magnitude <= RAW_VALUE_DECIMAL_MAX:
        return "dec", (f"The largest {noun} is {magnitude:.3g} — decimals. "
                       f"Read as percentages the fund never moved more than "
                       f"{magnitude:.3g} % in a period, which no real series "
                       f"sustains over a full history.")
    if magnitude < RAW_VALUE_TOO_QUIET:
        return None, (f"the series is too quiet to tell: the largest {noun} is "
                      f"{magnitude:.3g}, which is a believable fund as decimals "
                      f"({magnitude:.3g} per period) and as percentages "
                      f"({magnitude:.3g} % per period). Those two readings give "
                      f"SRRIs a hundredfold apart. State which it is.")
    return None, (f"percent and decimal cannot be separated: the largest {noun} "
                  f"is {magnitude:.3g}, plausible either way. State which it is.")


def detect_series_kind(values: Union[pd.Series, list, np.ndarray]) -> Detection:
    """Work out whether a column holds prices, periodic returns, or cumulative
    returns — and in percent or decimal.

    Returns `SeriesKind.AMBIGUOUS` with the shortlist attached rather than
    guessing when the evidence does not decide.
    """
    s = values if isinstance(values, pd.Series) else pd.Series(values)
    ev = _evidence(s)

    if ev.get("n", 0) < 3:
        return Detection(kind=SeriesKind.AMBIGUOUS, confident=False,
                         reason="Fewer than three usable values.", evidence=ev)

    # --- prices ------------------------------------------------------
    # A NAV cannot be negative, and a price series is a near-random walk in
    # levels: consecutive values sit close together, so lag-1 autocorrelation
    # of the LEVELS is very high and signs never flip.
    if (not ev["any_negative"] and ev["min"] > 0
            and ev["sign_flip_rate"] == 0.0
            and (ev["level_autocorr"] is None
                 or ev["level_autocorr"] >= LEVEL_AUTOCORR_PRICE_MIN)
            and abs(ev["first"]) > CUMULATIVE_FIRST_TOL):
        return Detection(
            kind=SeriesKind.PRICE, confident=True,
            reason=("All values positive, no sign changes, and levels are "
                    "highly autocorrelated — a price series."),
            evidence=ev)

    # --- cumulative --------------------------------------------------
    # Starts at zero and drifts a long way from its typical step size.
    starts_at_zero = abs(ev["first"]) <= CUMULATIVE_FIRST_TOL
    drifts = (ev["median_abs"] > 0
              and abs(ev["last"]) / ev["median_abs"] > CUMULATIVE_DRIFT_RATIO)
    if starts_at_zero and ev["sign_flip_rate"] < SIGN_FLIP_PERIODIC_MIN:
        # A cumulative series grows, so its own magnitude says nothing about
        # the units — a fund up 80 % reads 0.8 in decimals and 80 in percent,
        # but so does a fund up 0.8 % in percent. The STEP between consecutive
        # points is the period return, which is what carries the units.
        scale_kind, why = _percent_or_decimal(ev["max_abs_diff"], "step")
        if scale_kind is None:
            return Detection(
                kind=SeriesKind.AMBIGUOUS, confident=False,
                reason=f"Looks cumulative, but {why}",
                candidates=[SeriesKind.CUMULATIVE_PCT,
                            SeriesKind.CUMULATIVE_DECIMAL],
                evidence=ev)
        kind = (SeriesKind.CUMULATIVE_PCT if scale_kind == "pct"
                else SeriesKind.CUMULATIVE_DECIMAL)
        return Detection(
            kind=kind, confident=True,
            reason=("Starts at zero, rarely changes sign, and drifts well "
                    "beyond its typical step — cumulative return. " + why
                    if drifts else
                    "Starts at zero and rarely changes sign — cumulative "
                    "return. " + why),
            evidence=ev)

    # --- periodic returns -------------------------------------------
    looks_periodic = (ev["any_negative"]
                      or ev["sign_flip_rate"] >= SIGN_FLIP_PERIODIC_MIN)
    if looks_periodic:
        scale_kind, why = _percent_or_decimal(ev["max_abs"], "move")
        if scale_kind is None:
            return Detection(
                kind=SeriesKind.AMBIGUOUS, confident=False,
                reason=f"Periodic returns, but {why}",
                candidates=[SeriesKind.RETURN_PCT, SeriesKind.RETURN_DECIMAL],
                evidence=ev)
        kind = (SeriesKind.RETURN_PCT if scale_kind == "pct"
                else SeriesKind.RETURN_DECIMAL)
        return Detection(
            kind=kind, confident=True,
            reason=(f"Sign changes on {ev['sign_flip_rate']:.0%} of steps. "
                    + why),
            evidence=ev)

    return Detection(
        kind=SeriesKind.AMBIGUOUS, confident=False,
        reason=("The column matches neither a price series nor a return "
                "series on the usual markers."),
        evidence=ev)


def to_prices(values: Union[pd.Series, list, np.ndarray],
              index: Optional[pd.DatetimeIndex] = None,
              *,
              kind: Optional[SeriesKind] = None,
              base: float = 100.0,
              ) -> tuple[Optional[pd.Series], Detection, list[Finding]]:
    """Convert a supplied column into a price series.

    `kind` states what the column is and skips detection — which is what the
    form should do once the preparer has told us. Leave it None to detect.

    Returns `(prices, detection, findings)`. `prices` is None where the column
    is ambiguous and no `kind` was asserted; the findings say why, and the
    caller should stop rather than proceed on a guess.
    """
    findings: list[Finding] = []
    s = values if isinstance(values, pd.Series) else pd.Series(values)
    if index is not None:
        s = pd.Series(s.to_numpy(), index=index)

    detection = detect_series_kind(s)

    if kind is not None:
        if detection.kind is not kind and detection.confident:
            findings.append(Finding(
                code=FindingCode.SERIES_KIND_ASSERTED,
                severity=Severity.WARNING,
                message=(f"The column was declared as {kind.value}, but it reads "
                         f"as {detection.kind.value}. {detection.reason} The "
                         f"declaration was used."),
                remediation="Confirm the declaration before publishing.",
                detail={"declared": kind.value, "detected": detection.kind.value,
                        **detection.evidence}))
        else:
            findings.append(Finding(
                code=FindingCode.SERIES_KIND_ASSERTED,
                severity=Severity.INFO,
                message=f"The column was declared as {kind.value}.",
                detail={"declared": kind.value}))
        effective = kind
    else:
        if detection.kind is SeriesKind.AMBIGUOUS:
            findings.append(Finding(
                code=FindingCode.SERIES_KIND_AMBIGUOUS,
                severity=Severity.ERROR,
                message=(f"Cannot tell what this column holds. "
                         f"{detection.reason}"),
                remediation=("State the series type on the form. Guessing "
                             "between percent and decimal scales every figure "
                             "by 100."),
                detail={"candidates": [c.value for c in detection.candidates],
                        **detection.evidence}))
            return None, detection, findings
        findings.append(Finding(
            code=FindingCode.SERIES_KIND_DETECTED,
            severity=Severity.INFO,
            message=f"Column read as {detection.kind.value}. {detection.reason}",
            detail=detection.evidence))
        effective = detection.kind

    if effective is SeriesKind.PRICE:
        return pd.to_numeric(s, errors="coerce").dropna(), detection, findings

    numeric = pd.to_numeric(s, errors="coerce")
    scale = 100.0 if effective in (SeriesKind.RETURN_PCT,
                                   SeriesKind.CUMULATIVE_PCT) else 1.0

    if effective in (SeriesKind.CUMULATIVE_PCT, SeriesKind.CUMULATIVE_DECIMAL):
        prices = base * (1.0 + numeric / scale)
        prices = prices.dropna()
        note = (f"Cumulative returns rebased to {base:g}: "
                f"price = {base:g} x (1 + cumulative return).")
    else:
        r = numeric / scale
        leading_blank = bool(pd.isna(r.iloc[0])) if len(r) else False
        # The first row's return is the return of the period ENDING on the
        # first date, so the base sits one period earlier. Where the first row
        # is blank the series simply opens at the base.
        factors = (1.0 + r).fillna(1.0)
        prices = base * factors.cumprod()
        prices = prices[numeric.notna() | pd.Series(
            [i == 0 for i in range(len(numeric))], index=numeric.index)]
        note = (f"Periodic returns compounded from a base of {base:g}. "
                + ("The first row was blank, so the series opens at "
                   f"{base:g}."
                   if leading_blank else
                   "The first row's return is treated as the return of the "
                   f"period ending on the first date, so the {base:g} base "
                   "sits one period earlier."))

    # A blank FIRST row is expected, not a defect: the first period has no
    # return. Only interior blanks break the compounding chain.
    dropped = int(numeric.isna().sum())
    if len(numeric) and pd.isna(numeric.iloc[0]):
        dropped -= 1
    findings.append(Finding(
        code=FindingCode.RETURNS_CONVERTED_TO_PRICES,
        severity=Severity.INFO,
        message=(f"{note} The base is a convention — every KIID figure is a "
                 f"ratio of two points, so its value affects nothing."),
        detail={"kind": effective.value, "base": base, "scale_divisor": scale,
                "points_in": int(len(numeric)), "points_out": int(len(prices)),
                "unparseable_dropped": dropped}))

    if dropped:
        findings.append(Finding(
            code=FindingCode.PRICES_DROPPED,
            severity=Severity.WARNING,
            message=f"{dropped:,} return value(s) could not be read as numbers.",
            remediation="A gap in a return series silently shortens the "
                        "compounding chain; confirm the file is complete.",
            detail={"dropped": dropped}))

    return prices, detection, findings


def prices_from_upload(source: Union[str, bytes, "Path"],
                       *,
                       sheet: Optional[Union[str, int]] = None,
                       kind: Optional[SeriesKind] = None,
                       base: float = 100.0,
                       date_format=None,
                       ) -> tuple[Optional[pd.Series], list[Finding], UnitsHint]:
    """One call: uploaded workbook in, price series out.

    Handles both intakes without the caller having to know which arrived:

    - a NAV price column, returned unchanged;
    - a return column — percent-formatted cells, `"1.25%"` text, or a header
      naming a return — converted to prices based at `base`.

    Order of authority: an explicit `kind` beats what the file declares, and
    what the file declares beats the magnitude heuristics. The heuristics only
    run when the file says nothing, and they still refuse where percent and
    decimal cannot be separated.

    Feed the returned series straight to `past_performance.run_kiid()`.
    """
    from srri_engine import DateFormat, read_prices

    fmt = date_format if date_format is not None else DateFormat.ISO
    findings: list[Finding] = []

    hint = detect_units_from_workbook(source, sheet=sheet)
    if hint.kind is not None:
        findings.append(Finding(
            code=FindingCode.SERIES_KIND_DETECTED,
            severity=Severity.INFO,
            message=(f"The file declares its own units via {hint.source}: "
                     f"{hint.kind.value}. "
                     f"{hint.evidence.get('note', '')}"),
            detail=hint.evidence))

    effective = kind or hint.kind

    # A price column goes through `read_prices`, which is the parser both
    # engines already use.
    #
    # A RETURN column must NOT. `read_prices` is a price reader: it treats
    # zero and negative values as bad data and strips them. Half of any real
    # return series is negative, so routing returns through it silently deletes
    # every down period and collapses measured volatility. Returns are read
    # raw here and only become prices — positive by construction — afterwards.
    if effective is SeriesKind.PRICE or effective is None:
        parsed = read_prices(source, date_format=fmt, sheet=sheet)
        findings.extend(parsed.findings)
        if effective is SeriesKind.PRICE:
            return parsed.prices, findings, hint
        # Nothing declared — fall back to the magnitude heuristics.
        detected = detect_series_kind(parsed.prices)
        if detected.kind is SeriesKind.PRICE:
            return parsed.prices, findings, hint
        prices, _d, conv = to_prices(parsed.prices, parsed.prices.index,
                                     kind=None, base=base)
        findings.extend(conv)
        return prices, findings, hint

    raw = _read_raw_column(source, sheet=sheet, date_format=fmt)
    if raw is None or raw.empty:
        findings.append(Finding(
            code=FindingCode.NO_DATA, severity=Severity.ERROR,
            message="No usable rows were found under the header.",
            detail=hint.evidence))
        return None, findings, hint

    # Percent-formatted cells store the decimal already, so the decimal
    # readings are correct whatever the magnitudes suggest.
    if hint.source == "cell_format" and kind is None:
        effective = (SeriesKind.CUMULATIVE_DECIMAL
                     if effective is SeriesKind.CUMULATIVE_DECIMAL
                     else SeriesKind.RETURN_DECIMAL)

    prices, _detection, conv_findings = to_prices(raw, raw.index,
                                                  kind=effective, base=base)
    findings.extend(conv_findings)
    return prices, findings, hint


def _read_raw_column(source: Union[str, bytes, Path],
                     *,
                     sheet: Optional[Union[str, int]] = None,
                     date_format=None,
                     max_scan_rows: int = 60) -> Optional[pd.Series]:
    """Date-indexed values straight out of the workbook, uncleaned.

    Deliberately does no validity filtering. Negative values are the point of
    a return series, so nothing may drop them.
    """
    from srri_engine import DateFormat

    try:
        from openpyxl import load_workbook
        src = io.BytesIO(source) if isinstance(source, bytes) else str(source)
        wb = load_workbook(src, data_only=True)
        ws = wb[sheet] if isinstance(sheet, str) else (
            wb.worksheets[sheet] if isinstance(sheet, int) else wb.worksheets[0])
    except Exception:                                      # noqa: BLE001
        return None

    header_row = date_col = value_col = None
    for r in range(1, min(max_scan_rows, ws.max_row) + 1):
        texts = [(c, str(ws.cell(row=r, column=c).value))
                 for c in range(1, min(ws.max_column, 12) + 1)
                 if isinstance(ws.cell(row=r, column=c).value, str)
                 and str(ws.cell(row=r, column=c).value).strip()]
        if len(texts) < 2:
            continue
        below = [ws.cell(row=r + 1, column=c).value for c, _ in texts]
        if not any(_looks_numeric(v) for v in below):
            continue
        for c, text in texts:
            if "date" in text.lower():
                date_col = c
            elif value_col is None:
                value_col = c
        header_row = r
        break

    if header_row is None or value_col is None:
        return None
    if date_col is None:
        date_col = 1 if value_col != 1 else 2

    dates, values = [], []
    for r in range(header_row + 1, ws.max_row + 1):
        d = ws.cell(row=r, column=date_col).value
        v = ws.cell(row=r, column=value_col).value
        if d is None and v is None:
            continue
        dates.append(d)
        values.append(_to_number(v))

    fmt = date_format or DateFormat.ISO
    kwargs = {"dayfirst": fmt is DateFormat.DMY,
              "yearfirst": fmt is DateFormat.ISO}
    idx = pd.to_datetime(pd.Series(dates), errors="coerce", **kwargs)
    s = pd.Series(values, index=pd.DatetimeIndex(idx))
    s = s[~s.index.isna()].sort_index()
    return s[~s.index.duplicated(keep="last")]
