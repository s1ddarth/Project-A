"""test_nav_template.py — the NAV Request Template must parse cleanly.

We ship `frontend/public/nav-template.xlsx` and tell users to fill it in, so the
engine has to read it without complaint. That is a contract between two files in
different halves of the repo, and nothing else pins it.

The template puts ISIN / Fund Name / Period / Frequency / Source in rows 1-5,
the column headers on row 7, and the first observation on row 8. Read naively,
pandas takes row 1 as the header and the metadata block as data — which then
"works" only because the junk rows fail date parsing and get dropped. The engine
calls that a silent near-miss and has `_find_header_row` to prevent it.

It regressed once already: the template labels its price column "Daily NAV",
which was absent from `_PRICE_ALIASES`, so detection failed and every upload of
the house template produced a spurious DATES_DROPPED warning on a clean file.
"""
from __future__ import annotations

import sys
import warnings
from datetime import date
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")
sys.path.insert(0, str(Path(__file__).parent))

from srri_engine import (  # noqa: E402
    DateFormat, FindingCode, Severity, _find_header_row, _norm_header,
    _PRICE_ALIASES, _DATE_ALIASES, _sniff_kind, read_prices,
)

PASS, FAIL = [], []
TEMPLATE = (Path(__file__).resolve().parents[2]
            / "frontend" / "public" / "nav-template.xlsx")


def check(name, cond, detail=""):
    (PASS if cond else FAIL).append(name)
    print(f"  {'PASS' if cond else 'FAIL'}  {name}" + (f"  — {detail}" if detail and not cond else ""))


def filled_template(n_days: int = 1566) -> bytes:
    """The shipped template with realistic NAV data written into it."""
    from openpyxl import load_workbook
    import io

    wb = load_workbook(TEMPLATE)
    ws = wb[wb.sheetnames[0]]
    ws["B1"] = "IE00BDBB9Q16"
    ws["B2"] = "EPIC Financial Trends"
    ws["B4"] = "Daily"
    rng = np.random.default_rng(3)
    dates = pd.bdate_range("2019-01-01", periods=n_days)
    prices = 100 * np.exp(np.cumsum(rng.normal(0.0003, 0.08 / np.sqrt(252), n_days)))
    for i, (d, p) in enumerate(zip(dates, prices), start=8):
        ws.cell(i, 1, d.date())
        ws.cell(i, 2, round(float(p), 5))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


print("\nNAV Request Template — parse contract")

# ======================================================================
print("\n1. The shipped file is where the frontend expects it")

check("1.a template exists at frontend/public/nav-template.xlsx", TEMPLATE.exists(),
      f"looked in {TEMPLATE}")

if not TEMPLATE.exists():                              # nothing else can run
    print(f"\n  {len(PASS)} passed, {len(FAIL)} failed")
    sys.exit(1)

from openpyxl import load_workbook  # noqa: E402

ws = load_workbook(TEMPLATE)[load_workbook(TEMPLATE).sheetnames[0]]


def _row_norms(r):
    return {_norm_header(ws.cell(r, c).value) for c in range(1, 4)}


# The engine requires a date alias AND a price alias in the SAME row, which
# matters here: the template's metadata label "Period" is itself a date alias,
# so a date-only search picks row 3 and misses the real header on row 7.
header_row = next((r for r in range(1, 15)
                   if (_row_norms(r) & _DATE_ALIASES) and (_row_norms(r) & _PRICE_ALIASES)),
                  None)
check("1.b a row carrying BOTH a date and a price header exists",
      header_row is not None,
      "no row in the first 15 has both — _find_header_row cannot anchor")
check("1.c that row is the documented header row (7)",
      header_row == 7, f"found the header on row {header_row}, expected 7")
check("1.d 'Period' metadata does not masquerade as the header row",
      not (_row_norms(3) & _PRICE_ALIASES),
      "row 3 looks like a header to a date-only search")

# ======================================================================
print("\n2. Cadence-labelled price columns are recognised")

# The template's Frequency field offers daily/weekly/monthly, so all three
# labels must resolve — not just the one the current file happens to use.
for label in ("Daily NAV", "Weekly NAV", "Monthly NAV", "NAV", "Price"):
    check(f"2.a {label!r} resolves to a price column",
          _norm_header(label) in _PRICE_ALIASES)

# ======================================================================
print("\n3. The metadata block is skipped, not silently dropped")

raw = filled_template()
kind = _sniff_kind(raw, "nav.xlsx")
detected = _find_header_row(raw, kind, "nav.xlsx", None)
check("3.a header row is detected rather than guessed positionally",
      detected == 6, f"got {detected!r}, expected 6")

ps = read_prices(raw, date_format=DateFormat.ISO, filename="nav.xlsx")
check("3.b every data row survives — none dropped as unparseable",
      ps.raw_rows == len(ps.prices) == 1566,
      f"raw_rows={ps.raw_rows} kept={len(ps.prices)}")

dropped = [f for f in ps.findings if f.code is FindingCode.DATES_DROPPED]
check("3.c a clean template produces no DATES_DROPPED warning",
      not dropped, dropped[0].message if dropped else "")

check("3.d no warnings at all on a clean template",
      not [f for f in ps.findings if f.severity is Severity.WARNING],
      str([f.code.value for f in ps.findings if f.severity is Severity.WARNING]))

check("3.e the detection is reported to the user, not silent",
      any(f.code is FindingCode.COLUMNS_ASSUMED for f in ps.findings))

check("3.f dates parse to the intended range",
      ps.prices.index[0].date() == date(2019, 1, 1))

# ======================================================================
print("\n4. End to end — the shipped template yields a document")

from past_performance import run_kiid  # noqa: E402

srri, pp = run_kiid(raw, reference_date=date(2024, 12, 31), currency="USD",
                    date_format=DateFormat.ISO, filename="nav-template.xlsx")
check("4.a SRRI is produced from the house template", srri.srri_disclosed is not None,
      f"status {srri.status.value}")
check("4.b past performance is produced too", len(pp.bars) > 0,
      f"status {pp.status.value}")
check("4.c both halves share one input hash",
      srri.audit.input_sha256 == pp.audit.input_sha256)

# ======================================================================
print("\n" + "=" * 70)
print(f"  {len(PASS)} passed, {len(FAIL)} failed")
if FAIL:
    for f in FAIL:
        print(f"    FAILED: {f}")
print("=" * 70)
sys.exit(1 if FAIL else 0)
