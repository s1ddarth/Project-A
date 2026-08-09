"""
srri_engine.py — UCITS KIID SRRI calculation engine (CESR/10-673)
==================================================================

Single engine for BOTH the weekly (m=52, T=260) and monthly (m=12, T=60)
bases of CESR/10-673 Box 1 §4.  Replaces the previous pair of scripts
(`SRRI_Weekly_Calculator_Daily_to_Weekly.py`, `SRRI_Monthly_Calculator.py`),
which were ~85% duplicated code — the buffer-zone logic in particular existed
in two places and had already diverged (see NOTE 3 below).

Design contract (web-service first, CLI second)
-----------------------------------------------
1. Frequency is a parameter, not a module.        -> `Frequency.WEEKLY|MONTHLY|AUTO`
2. Input is bytes / file object / path / DataFrame. -> `read_prices(...)`
3. The core returns a result object, never a file.   -> `calculate(...) -> SRRIResult`
4. Excel export is an optional function on a result. -> `export_workbook(result)`
5. Data-quality checks are a separate step returning
   tagged findings the UI can render.                -> `validate(...) -> list[Finding]`
6. No argparse in the core.                          -> `main()` is a thin wrapper.
7. Nothing is reported by logging alone.             -> every condition is a `Finding`.
8. Date format is an explicit input; the sub-5-year
   window can only be relaxed by an attributable,
   logged override.                                  -> `DateFormat`, `MinPeriodsOverride`
9. Every result carries the engine version and a
   SHA-256 of the exact bytes parsed.                -> `AuditInfo`

Usage (library)
---------------
    from srri_engine import Frequency, DateFormat, run, export_workbook

    result = run(uploaded_bytes, frequency=Frequency.WEEKLY,
                 date_format=DateFormat.DMY, filename="nav.xlsx")

    if result.is_blocked:
        return {"errors": [f.model_dump() for f in result.errors]}

    xlsx_bytes = export_workbook(result)          # audit artifact
    payload    = result.model_dump(mode="json")   # persist this

Usage (CLI — thin wrapper, not the entry point)
-----------------------------------------------
    python srri_engine.py nav.xlsx --frequency weekly --date-format dmy \
        --output srri.xlsx

Regulatory basis: CESR/10-673, 1 July 2010 (Box 1 §2/§4, Box 2, Box 3);
Commission Regulation (EU) No 583/2010.

NOTE 1 (bug fixed): `dayfirst=True` was hardcoded in both scripts, so
  "03/04/2024" silently parsed as 3 April.  Date format is now an explicit,
  required-by-signature input, and genuinely ambiguous files raise a finding.
NOTE 2 (bug fixed): `--min-weeks` / `--min-months` allowed an SRRI to be
  produced on under 5 years of data from an unattributed CLI flag.  Relaxation
  now requires a `MinPeriodsOverride` carrying approver + reason, which is
  echoed into the result and the workbook.
NOTE 3 (bug fixed): the monthly script applied the Box 3 buffer over the
  leading NaN ("Insufficient Data") rows.  Because `NaN == x` is False, the
  first disclosed SRRI was suppressed for `buffer_months - 1` extra months
  relative to the weekly script.  Both bases now condense to monthly reference
  points, drop NaN, buffer, and forward-fill — one code path, no divergence.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import logging
import re
import sys
from datetime import date, datetime, timezone
from enum import Enum
from pathlib import Path
from typing import IO, Any, Optional, Union

import numpy as np
import pandas as pd
from pydantic import BaseModel, ConfigDict, Field, field_validator

log = logging.getLogger("srri_engine")

# ======================================================================
# Identity — every result is traceable to exactly this code
# ======================================================================

ENGINE_NAME = "srri_engine"
ENGINE_VERSION = "3.0.0"   # 3.0.0: CESR Box 3 §2/§3 corrected — disclosed
                           # SRRI values move, so this is a MAJOR bump.
                           # A stored engine_version must identify exactly
                           # one behaviour (rule 5).
METHODOLOGY_REF = (
    "CESR/10-673 (1 July 2010); Commission Regulation (EU) No 583/2010"
)

# ======================================================================
# Regulatory constants
# ======================================================================

# Box 2 — volatility bands (lower inclusive, upper exclusive)
SRRI_BANDS: list[tuple[int, float, float]] = [
    (1, 0.000, 0.005),          # 0.00%  – 0.50%
    (2, 0.005, 0.020),          # 0.50%  – 2.00%
    (3, 0.020, 0.050),          # 2.00%  – 5.00%
    (4, 0.050, 0.100),          # 5.00%  – 10.00%
    (5, 0.100, 0.150),          # 10.00% – 15.00%
    (6, 0.150, 0.250),          # 15.00% – 25.00%
    (7, 0.250, float("inf")),   # >= 25.00%
]

RISK_LABELS: dict[int, str] = {
    1: "Lowest Risk",       2: "Very Low Risk",   3: "Low Risk",
    4: "Medium Risk",       5: "Medium-High Risk",
    6: "High Risk",         7: "Highest Risk",
}

BUFFER_MONTHS_DEFAULT = 4      # Box 3 migration rule

# Input-cadence detection (median calendar days between observations)
DAILY_MAX_GAP = 2.5
WEEKLY_MIN_GAP, WEEKLY_MAX_GAP = 4.0, 9.0
MONTHLY_MIN_GAP, MONTHLY_MAX_GAP = 25.0, 45.0


class Frequency(str, Enum):
    """Calculation basis.  AUTO resolves from the input cadence, preferring
    weekly per CESR Box 1 §2 whenever the input supports it."""
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    AUTO = "auto"


class InputCadence(str, Enum):
    DAILY = "daily"
    WEEKLY = "weekly"
    MONTHLY = "monthly"
    IRREGULAR = "irregular"


class FrequencySpec(BaseModel):
    """The only thing that differs between the two regulatory bases."""
    model_config = ConfigDict(frozen=True)

    frequency: Frequency
    m: int                      # periods per year
    window: int                 # T — observations in the 5-year window
    resample_rule: str          # pandas rule used to build the period grid
    period_noun: str            # "week" / "month"
    period_noun_plural: str
    basis_label: str

    @property
    def annualisation(self) -> str:
        return f"sqrt({self.m})"


FREQUENCY_SPEC: dict[Frequency, FrequencySpec] = {
    Frequency.WEEKLY: FrequencySpec(
        frequency=Frequency.WEEKLY, m=52, window=260, resample_rule="W-FRI",
        period_noun="week", period_noun_plural="weeks",
        basis_label="Weekly (Friday close)",
    ),
    Frequency.MONTHLY: FrequencySpec(
        frequency=Frequency.MONTHLY, m=12, window=60, resample_rule="ME",
        period_noun="month", period_noun_plural="months",
        basis_label="Monthly (month-end)",
    ),
}


class DateFormat(str, Enum):
    """Explicit day/month ordering.  Replaces the hardcoded `dayfirst=True`.

    OPEN POINT (2026-08-07): the house default is DMY, matching the European
    NAV files we currently receive.  Before the uploader is exposed to
    external users this should become a required, per-upload choice with no
    default — a US-formatted file silently parsed as DMY produces a wrong
    SRRI with no other symptom.  `AUTO` and the ambiguity findings below are
    already in place to support that switch.
    """
    DMY = "dmy"      # 03/04/2024 -> 3 April 2024
    MDY = "mdy"      # 03/04/2024 -> 4 March 2024
    ISO = "iso"      # 2024-04-03
    AUTO = "auto"    # infer; errors out when genuinely ambiguous


# ======================================================================
# Findings — every condition the user must see is a returned value
# ======================================================================

class Severity(str, Enum):
    ERROR = "error"       # blocks the user
    WARNING = "warning"   # user may proceed after acknowledging
    INFO = "info"         # displayed, no acknowledgement required


class FindingCode(str, Enum):
    """Stable codes.  The UI keys off these, not off the message text."""
    # --- errors
    NO_DATA = "NO_DATA"
    UNREADABLE_INPUT = "UNREADABLE_INPUT"
    NO_DATE_COLUMN = "NO_DATE_COLUMN"
    NO_PRICE_COLUMN = "NO_PRICE_COLUMN"
    ALL_DATES_UNPARSEABLE = "ALL_DATES_UNPARSEABLE"
    ALL_PRICES_UNPARSEABLE = "ALL_PRICES_UNPARSEABLE"
    TOO_FEW_OBSERVATIONS = "TOO_FEW_OBSERVATIONS"
    DATE_FORMAT_AMBIGUOUS = "DATE_FORMAT_AMBIGUOUS"
    DATE_FORMAT_CONFLICT = "DATE_FORMAT_CONFLICT"
    CADENCE_TOO_COARSE = "CADENCE_TOO_COARSE"
    CADENCE_IRREGULAR = "CADENCE_IRREGULAR"
    INSUFFICIENT_HISTORY = "INSUFFICIENT_HISTORY"
    # --- warnings
    SHORT_HISTORY = "SHORT_HISTORY"
    DATES_DROPPED = "DATES_DROPPED"
    PRICES_DROPPED = "PRICES_DROPPED"
    NON_POSITIVE_PRICES = "NON_POSITIVE_PRICES"
    DUPLICATE_DATES = "DUPLICATE_DATES"
    UNSORTED_DATES = "UNSORTED_DATES"
    LARGE_GAP = "LARGE_GAP"
    STALE_PRICES = "STALE_PRICES"
    EXTREME_RETURN = "EXTREME_RETURN"
    MONTHLY_BASIS_NOT_PREFERRED = "MONTHLY_BASIS_NOT_PREFERRED"
    MIN_PERIODS_OVERRIDE_APPLIED = "MIN_PERIODS_OVERRIDE_APPLIED"
    DATE_FORMAT_ASSUMED = "DATE_FORMAT_ASSUMED"
    # --- info
    COLUMNS_ASSUMED = "COLUMNS_ASSUMED"
    CADENCE_DETECTED = "CADENCE_DETECTED"
    RESAMPLED = "RESAMPLED"
    FREQUENCY_AUTO_SELECTED = "FREQUENCY_AUTO_SELECTED"
    DATE_FORMAT_INFERRED = "DATE_FORMAT_INFERRED"

    # ------------------------------------------------------------------
    # Upstream validation codes — raised by the pre-parse validation step,
    # not by this module, and passed in via `extra_findings` so they surface
    # in the same findings list and the same audit sheets as everything else.
    # ------------------------------------------------------------------
    CURRENCY_MISMATCH = "CURRENCY_MISMATCH"
    TEMPLATE_CONTRACT_BREACH = "TEMPLATE_CONTRACT_BREACH"
    UPSTREAM_VALIDATION = "UPSTREAM_VALIDATION"
    # Return-series intake — see `returns_input.py`
    SERIES_KIND_DETECTED = "SERIES_KIND_DETECTED"
    SERIES_KIND_AMBIGUOUS = "SERIES_KIND_AMBIGUOUS"
    SERIES_KIND_ASSERTED = "SERIES_KIND_ASSERTED"
    RETURNS_CONVERTED_TO_PRICES = "RETURNS_CONVERTED_TO_PRICES"

    # ------------------------------------------------------------------
    # Past-performance codes — raised by `past_performance.py`
    # (Reg. (EU) No 583/2010, Arts. 15-19 and Annex III), not by this module.
    #
    # They live here because a KIID is one document produced from one NAV
    # upload: the SRRI and the bar chart are two readings of the same series,
    # and the UI renders one findings list.  Two `Finding` classes with two
    # code enums would force every consumer to branch on which engine spoke.
    # Adding members is additive and breaks nothing that keys off the
    # existing codes.
    # ------------------------------------------------------------------
    # errors
    NO_COMPLETE_CALENDAR_YEAR = "NO_COMPLETE_CALENDAR_YEAR"        # Art. 15(4)
    REFERENCE_DATE_BEFORE_SERIES = "REFERENCE_DATE_BEFORE_SERIES"
    BENCHMARK_UNUSABLE = "BENCHMARK_UNUSABLE"
    # warnings
    FEWER_THAN_FIVE_YEARS = "FEWER_THAN_FIVE_YEARS"                # Art. 15(2)
    MISSING_YEAR = "MISSING_YEAR"                                  # Art. 15(3)
    STALE_YEAR_ANCHOR = "STALE_YEAR_ANCHOR"
    EXTREME_ANNUAL_RETURN = "EXTREME_ANNUAL_RETURN"
    BENCHMARK_YEAR_MISSING = "BENCHMARK_YEAR_MISSING"              # Art. 18(2)
    CLIENT_FIGURE_MISMATCH = "CLIENT_FIGURE_MISMATCH"
    CLIENT_FIGURE_YEAR_UNMATCHED = "CLIENT_FIGURE_YEAR_UNMATCHED"
    MATERIAL_CHANGE_IN_WINDOW = "MATERIAL_CHANGE_IN_WINDOW"        # Art. 17
    SIMULATED_PERFORMANCE_SHOWN = "SIMULATED_PERFORMANCE_SHOWN"    # Art. 19(2)
    PARTIAL_YEAR_EXCLUDED = "PARTIAL_YEAR_EXCLUDED"
    STALE_PUBLICATION_WINDOW = "STALE_PUBLICATION_WINDOW"          # Art. 23(3)
    # info
    CURRENT_YEAR_EXCLUDED = "CURRENT_YEAR_EXCLUDED"                # Art. 15(6)
    CHART_WINDOW_SELECTED = "CHART_WINDOW_SELECTED"
    ASSUMED_NET_OF_ONGOING_CHARGES = "ASSUMED_NET_OF_ONGOING_CHARGES"
    ASSUMED_DISTRIBUTION_ADJUSTED = "ASSUMED_DISTRIBUTION_ADJUSTED"
    ANCHOR_TOLERANCE_SET = "ANCHOR_TOLERANCE_SET"
    CLIENT_FIGURES_RECONCILED = "CLIENT_FIGURES_RECONCILED"


class Finding(BaseModel):
    """One data-quality or methodology observation, renderable by the UI."""
    model_config = ConfigDict(frozen=True)

    code: FindingCode
    severity: Severity
    message: str
    detail: dict[str, Any] = Field(default_factory=dict)
    remediation: Optional[str] = None

    def __str__(self) -> str:                                  # pragma: no cover
        return f"[{self.severity.value.upper()}] {self.code.value}: {self.message}"


def _err(code, msg, remediation=None, **detail) -> Finding:
    return Finding(code=code, severity=Severity.ERROR, message=msg,
                   remediation=remediation, detail=detail)


def _warn(code, msg, remediation=None, **detail) -> Finding:
    return Finding(code=code, severity=Severity.WARNING, message=msg,
                   remediation=remediation, detail=detail)


def _info(code, msg, **detail) -> Finding:
    return Finding(code=code, severity=Severity.INFO, message=msg, detail=detail)


class SRRIInputError(ValueError):
    """Raised only when the input cannot be turned into a price series at all.

    Everything recoverable is a `Finding`, not an exception — the web layer
    should be able to show the user a list, not a stack trace.
    """

    def __init__(self, finding: Finding):
        self.finding = finding
        super().__init__(finding.message)


# ======================================================================
# The override — sub-5-year windows must be attributable
# ======================================================================

class MinPeriodsOverride(BaseModel):
    """Admin-only relaxation of the CESR 5-year window.

    The engine will not shorten the window without one of these, and it
    cannot be constructed without an approver and a reason.  The whole
    object is echoed into `SRRIResult.audit` and onto the workbook's
    Audit & Findings sheet, so a published document can always answer
    "who allowed this to be produced on 3.2 years of data?".
    """
    model_config = ConfigDict(frozen=True)

    min_periods: int = Field(..., gt=1,
                             description="Relaxed minimum observations in the window.")
    approved_by: str = Field(..., min_length=2,
                             description="Named individual, not a role or a system account.")
    reason: str = Field(..., min_length=10,
                        description="Why a non-standard window is acceptable for this fund.")
    approved_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    ticket_reference: Optional[str] = None

    @field_validator("approved_by", "reason")
    @classmethod
    def _not_placeholder(cls, v: str) -> str:
        s = v.strip()
        if s.lower() in {"n/a", "na", "none", "tbc", "tbd", "-", "test", "admin", "user"}:
            raise ValueError("approved_by / reason must be a real, attributable value")
        return s


# ======================================================================
# Result models
# ======================================================================

class ResultStatus(str, Enum):
    OK = "ok"                            # valid SRRI, no warnings
    OK_WITH_WARNINGS = "ok_with_warnings"  # valid SRRI, user must acknowledge
    NO_VALID_SRRI = "no_valid_srri"      # ran, but never met the window
    BLOCKED = "blocked"                  # errors — nothing was calculated


class AuditInfo(BaseModel):
    """Everything needed to answer "why did the March document say SRRI 4?"
    six months later, from stored data alone."""
    model_config = ConfigDict(frozen=True)

    engine_name: str = ENGINE_NAME
    engine_version: str = ENGINE_VERSION
    methodology_ref: str = METHODOLOGY_REF
    calculated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

    input_sha256: str
    input_filename: Optional[str] = None
    input_bytes: int = 0
    input_kind: str = "bytes"            # bytes | path | fileobj | dataframe | series

    # resolved parameters actually used (not what was requested)
    frequency: Frequency
    m: int
    window: int
    annualisation: str
    date_format: DateFormat
    date_format_resolved: DateFormat
    buffer_months: int
    min_periods: int
    min_periods_is_regulatory_default: bool
    min_periods_override: Optional[MinPeriodsOverride] = None
    sheet: Optional[Union[str, int]] = None
    skiprows: int = 0

    @property
    def fingerprint(self) -> str:
        """Short, stable handle to quote on a document or in a ticket."""
        return f"{self.engine_version}/{self.input_sha256[:12]}"


class SeriesPoint(BaseModel):
    """One reference point on the calculation grid."""
    model_config = ConfigDict(frozen=True)

    date: date
    price: float
    period_return_pct: Optional[float] = None
    periods_in_window: int = 0
    mean_return_pct: Optional[float] = None
    period_vol_pct: Optional[float] = None
    ann_vol_pct: Optional[float] = None
    srri_raw: Optional[int] = None
    srri_disclosed: Optional[int] = None
    risk_description: str = "Insufficient Data"
    periods_at_srri: int = 0
    status: str = "Insufficient Data"


class SRRIResult(BaseModel):
    """The return value of the core.  Excel is derived from this, never the
    other way round."""
    model_config = ConfigDict(frozen=True)

    status: ResultStatus
    audit: AuditInfo
    findings: list[Finding] = Field(default_factory=list)

    # headline
    as_of_date: Optional[date] = None
    annualised_volatility: Optional[float] = None      # decimal, e.g. 0.1234
    srri_raw: Optional[int] = None
    srri_disclosed: Optional[int] = None
    risk_description: Optional[str] = None

    # provenance of the series
    input_cadence: Optional[InputCadence] = None
    input_first_date: Optional[date] = None
    input_last_date: Optional[date] = None
    input_rows: int = 0
    history_years: Optional[float] = None

    # calculation grid
    n_periods: int = 0
    n_valid_periods: int = 0
    first_valid_date: Optional[date] = None
    series: list[SeriesPoint] = Field(default_factory=list)
    distribution: dict[int, int] = Field(default_factory=dict)

    # ------------------------------------------------------------------
    @property
    def errors(self) -> list[Finding]:
        return [f for f in self.findings if f.severity is Severity.ERROR]

    @property
    def warnings(self) -> list[Finding]:
        return [f for f in self.findings if f.severity is Severity.WARNING]

    @property
    def infos(self) -> list[Finding]:
        return [f for f in self.findings if f.severity is Severity.INFO]

    @property
    def is_blocked(self) -> bool:
        return self.status is ResultStatus.BLOCKED

    @property
    def requires_acknowledgement(self) -> bool:
        return bool(self.warnings)

    @property
    def annualised_volatility_pct(self) -> Optional[float]:
        return None if self.annualised_volatility is None else self.annualised_volatility * 100

    def to_frame(self) -> pd.DataFrame:
        """The calculation grid as a DataFrame (for the workbook and for tests)."""
        if not self.series:
            return pd.DataFrame()
        df = pd.DataFrame([p.model_dump() for p in self.series])
        return df.set_index(pd.DatetimeIndex(df.pop("date"), name="Date"))

    def summary(self) -> str:
        a = self.audit
        lines = [
            f"SRRI {a.frequency.value.upper()} — CESR/10-673 "
            f"(m={a.m}, T={a.window}, {a.annualisation})",
            f"  engine {a.engine_version} | input {a.input_sha256[:12]} | status {self.status.value}",
        ]
        if self.srri_disclosed is not None:
            lines.append(
                f"  as of {self.as_of_date} | ann. vol {self.annualised_volatility_pct:.2f}% "
                f"| SRRI raw {self.srri_raw} | SRRI disclosed {self.srri_disclosed} "
                f"({self.risk_description})"
            )
        for f in self.findings:
            lines.append(f"  {f}")
        return "\n".join(lines)


# ======================================================================
# Core maths — one implementation, both bases
# ======================================================================

def classify_srri(ann_vol: float) -> Optional[int]:
    """Box 2 — volatility to SRRI class."""
    if ann_vol is None or pd.isna(ann_vol):
        return None
    for srri, lo, hi in SRRI_BANDS:
        if lo <= ann_vol < hi:
            return srri
    return 7


def cesr_volatility(returns, m: int) -> float:
    """Box 1 §4, written literally:  sigma = sqrt[ (m/(T-1)) * SUM (r_t - r_bar)^2 ].

    Kept as the reference implementation.  `calculate()` uses the algebraically
    identical rolling `std(ddof=1) * sqrt(m)` for speed and cross-checks a
    sample of points against this function (`verify_parity=True`).
    """
    r = np.asarray(returns, dtype=float)
    r = r[~np.isnan(r)]
    T = len(r)
    if T < 2:
        return float("nan")
    r_bar = r.mean()                                 # arithmetic mean (Box 1 §4)
    sum_sq = float(np.sum((r - r_bar) ** 2))         # SUM (r_t - r_bar)^2
    return float(np.sqrt((m / (T - 1)) * sum_sq))    # sqrt( (m/(T-1)) * SUM )


def _buffer_core(s: pd.Series, buffer_months: int) -> pd.Series:
    """Box 3 applied to a clean, NaN-free series of monthly reference points.

    Two stages, exactly as CESR/10-673 Box 3 defines them:

    §2 — TRIGGER. The indicator is revised only if the volatility "has fallen
         outside the bucket corresponding to its previous risk category on each
         weekly or monthly data reference point over the preceding 4 months".
         The test is therefore about *departure from the currently disclosed
         class*, not about arrival at a single new one. One reading back at the
         old class resets the count.

    §3 — SELECTION. Once triggered, and where the volatility "has moved so as to
         correspond to more than one bucket during the 4-month period", the fund
         takes "the bucket which its relevant volatility has matched for the
         majority of the ... data reference point[s]".

    NOTE: the earlier implementation required every point in the window to equal
    the *same* new class (`all(win == cur)`). That condition is strictly stronger
    than §2, so it never migrated when volatility straddled two new buckets — a
    fund could hold a stale class indefinitely, understating risk when volatility
    had risen and overstating it when it had fallen. §3 was absent entirely.

    HOUSE CONVENTION: §3 says "majority" but is silent on ties (e.g. 5,6,5,6 over
    four points). Ties resolve to the HIGHER risk class — where the rule does not
    decide, disclose the greater risk.

    The window is a real 4-CALENDAR-MONTH lookback, not a fixed count of points,
    so "each ... data reference point over the preceding 4 months" means every
    point on the calculation grid: ~17 weekly points on a weekly basis, 4 monthly
    points on a monthly basis. See `apply_buffer_zone`.
    """
    vals = s.to_numpy(dtype=float)
    idx = s.index
    out = np.full(len(vals), np.nan)
    for i, cur in enumerate(vals):
        if i == 0:
            out[i] = cur
            continue
        prev = out[i - 1]
        if cur == prev:
            out[i] = prev
            continue

        window_start = idx[i] - pd.DateOffset(months=buffer_months)
        # Month-end reference points are not evenly spaced (30 vs 31 days), so a
        # plain month offset from e.g. 30 Nov lands on 30 Jul and drags the 31 Jul
        # point into the window — 5 monthly points instead of 4. Snap the anchor
        # back onto month-end when the reference point itself is a month-end.
        if idx[i].is_month_end:
            window_start = window_start + pd.offsets.MonthEnd(0)

        # §2 — the series must actually span the full 4 months before a
        # migration can be confirmed at all.
        if idx[0] > window_start:
            out[i] = prev
            continue

        win = vals[idx.searchsorted(window_start, side="right"): i + 1]

        # §2 — the old class reappeared at some reference point: no revision.
        if len(win) == 0 or bool(np.any(win == prev)):
            out[i] = prev
            continue

        # §3 — majority bucket across the window; ties go to the higher class.
        classes, counts = np.unique(win, return_counts=True)
        out[i] = float(classes[counts == counts.max()].max())
    return pd.Series(out, index=s.index)


def apply_buffer_zone(srri_raw: pd.Series,
                      buffer_months: int = BUFFER_MONTHS_DEFAULT) -> pd.Series:
    """Box 3 buffer zone — THE single implementation, used by both bases.

    Box 3 §2 tests "each weekly OR monthly data reference point over the
    preceding 4 months".  The reference points are those of the basis the fund
    is calculated on, so the rule is applied directly to the calculation grid:

        weekly basis  -> ~17 weekly reference points in the 4-month window
        monthly basis -> 4 monthly reference points in the 4-month window

    HOUSE DECISION (2026-08-09): monitoring frequency follows the calculation
    basis.  A previous version condensed every series to month-ends with
    `resample("ME")` before applying the rule, so a weekly-basis fund was tested
    on 4 month-end snapshots rather than every week.  That let a fund migrate on
    four month-end readings while falling back to the old class in the weeks
    between.  Both readings are permitted by the wording; this is the stricter
    one and the one that matches how the fund is actually calculated.

    Dropping NaN before buffering is what fixes NOTE 3 in the module docstring.
    """
    if srri_raw.empty:
        return srri_raw.astype(float)
    points = srri_raw.dropna()
    if points.empty:
        return pd.Series(np.nan, index=srri_raw.index)
    disclosed = _buffer_core(points, buffer_months)
    return disclosed.reindex(srri_raw.index, method="ffill")


# ======================================================================
# Step 1 — read (bytes / file object / path / DataFrame / Series)
# ======================================================================

Source = Union[str, Path, bytes, bytearray, memoryview, IO[bytes], pd.DataFrame, pd.Series]

_DATE_ALIASES = {
    "date", "dates", "valuation date", "valuationdate", "nav date", "navdate",
    "week", "week ending", "week-ending", "weekending",
    "month", "month end", "month-end", "monthend",
    "day", "time", "datetime", "period", "as of", "as-of", "asof",
}
_PRICE_ALIASES = {
    "price", "prices", "nav", "nav price", "navprice", "nav per share",
    "close", "closing price", "adj close", "adjusted close",
    "value", "level", "unit price", "share price", "mid", "bid",
    # The house NAV Request Template labels the column by its cadence, and its
    # Frequency field implies all three. Without these, _find_header_row cannot
    # match the header and the metadata block above it is read as data — the
    # "silent near-miss" that function exists to prevent.
    "daily nav", "weekly nav", "monthly nav",
}

_DMY_LIKE = re.compile(r"^\s*(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})")
_ISO_LIKE = re.compile(r"^\s*(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})")


class PriceSeries(BaseModel):
    """Parsed, cleaned input plus the findings produced while parsing it."""
    model_config = ConfigDict(frozen=True, arbitrary_types_allowed=True)

    prices: pd.Series                # float values, DatetimeIndex, sorted, unique
    findings: list[Finding] = Field(default_factory=list)
    sha256: str
    filename: Optional[str] = None
    n_bytes: int = 0
    input_kind: str = "bytes"
    raw_rows: int = 0
    date_format_requested: DateFormat = DateFormat.DMY
    date_format_resolved: DateFormat = DateFormat.DMY
    date_column: Optional[str] = None
    price_column: Optional[str] = None


def _materialise(source: Source, filename: Optional[str]) -> tuple[bytes, str, str, Optional[str]]:
    """Normalise any accepted source to (raw_bytes, sha256, kind, filename).

    Hashing the exact bytes that get parsed — rather than re-reading a path
    later — is what makes the hash a real provenance record.
    """
    if isinstance(source, (pd.DataFrame, pd.Series)):
        kind = "dataframe" if isinstance(source, pd.DataFrame) else "series"
        frame = source.to_frame() if isinstance(source, pd.Series) else source
        raw = frame.to_csv(index=True).encode("utf-8")
        return raw, hashlib.sha256(raw).hexdigest(), kind, filename

    if isinstance(source, (bytes, bytearray, memoryview)):
        raw = bytes(source)
        return raw, hashlib.sha256(raw).hexdigest(), "bytes", filename

    if isinstance(source, (str, Path)):
        p = Path(source)
        if not p.exists():
            raise SRRIInputError(_err(FindingCode.UNREADABLE_INPUT,
                                      f"File not found: {p}",
                                      "Check the path, or upload the file instead."))
        raw = p.read_bytes()
        return raw, hashlib.sha256(raw).hexdigest(), "path", filename or p.name

    if hasattr(source, "read"):
        try:
            if hasattr(source, "seek"):
                source.seek(0)
            data = source.read()
        except Exception as exc:                                # pragma: no cover
            raise SRRIInputError(_err(FindingCode.UNREADABLE_INPUT,
                                      f"Could not read the uploaded file object: {exc}")) from exc
        raw = data.encode("utf-8") if isinstance(data, str) else bytes(data)
        name = filename or getattr(source, "name", None)
        name = Path(name).name if isinstance(name, str) else None
        return raw, hashlib.sha256(raw).hexdigest(), "fileobj", name

    raise SRRIInputError(_err(FindingCode.UNREADABLE_INPUT,
                              f"Unsupported input type: {type(source).__name__}",
                              "Pass bytes, a file object, a path, or a pandas object."))


def _sniff_kind(raw: bytes, filename: Optional[str]) -> str:
    """Excel vs delimited text, from magic bytes first, extension second."""
    if raw[:4] == b"PK\x03\x04":          # xlsx/xlsm (zip)
        return "excel"
    if raw[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":   # legacy xls (OLE2)
        return "excel"
    if filename:
        sfx = Path(filename).suffix.lower()
        if sfx in (".xlsx", ".xlsm", ".xls", ".xltx"):
            return "excel"
        if sfx in (".csv", ".txt", ".tsv", ".dat"):
            return "text"
    return "text"


def _read_tabular(raw: bytes, kind: str, filename: Optional[str],
                  sheet: Optional[Union[str, int]], skiprows: int) -> pd.DataFrame:
    buf = io.BytesIO(raw)
    if kind == "excel":
        try:
            df = pd.read_excel(buf, sheet_name=sheet if sheet is not None else 0,
                               skiprows=skiprows)
        except Exception as exc:
            raise SRRIInputError(_err(FindingCode.UNREADABLE_INPUT,
                                      f"Could not read the workbook: {exc}",
                                      "Check the sheet name and the number of header rows to skip.",
                                      sheet=sheet, skiprows=skiprows)) from exc
        if isinstance(df, dict):                       # sheet_name=None
            df = next(iter(df.values()))
        return df

    last_exc: Optional[Exception] = None
    for sep in (None, ",", "\t", ";", r"\s+"):
        try:
            buf.seek(0)
            df = pd.read_csv(buf, sep=sep, skiprows=skiprows, engine="python")
            if df.shape[1] >= 2:
                return df
        except Exception as exc:
            last_exc = exc
    raise SRRIInputError(_err(FindingCode.UNREADABLE_INPUT,
                              f"Could not parse {filename or 'the input'} as delimited text"
                              + (f" ({last_exc})" if last_exc else ""),
                              "Save the file as CSV or XLSX with a date column and a price column."))


_PAREN_SUFFIX = re.compile(r"\s*\([^)]*\)\s*$")


def _norm_header(name: object) -> str:
    """Normalise a column label for alias lookup.

    Strips a trailing parenthetical so that 'NAV Price (USD)' — the header the
    KIID NAV template writes, with the currency appended — still matches the
    'nav price' alias instead of falling through to positional guessing.
    """
    s = str(name).strip().lower()
    return _PAREN_SUFFIX.sub("", s).strip()


def _find_header_row(raw: bytes, kind: str, filename: Optional[str],
                     sheet: Optional[Union[str, int]], scan: int = 15) -> Optional[int]:
    """Locate the real header row when a file carries a metadata block above it.

    The KIID NAV template puts ISIN / Fund Name / Period / Frequency / Source in
    rows 1-5, the column headers on row 7 and the first observation on row 8.
    Read naively, pandas takes row 1 as the header and the metadata block as
    data — which then survives only because the junk rows fail date parsing and
    get dropped. That is a silent near-miss, so find the header row properly.

    Returns the number of rows to skip, or None if no better row exists.
    """
    try:
        probe = _read_tabular(raw, kind, filename, sheet, 0)
    except SRRIInputError:
        return None

    def row_matches(values) -> bool:
        norm = {_norm_header(v) for v in values if v is not None and str(v) != "nan"}
        return bool(norm & _DATE_ALIASES) and bool(norm & _PRICE_ALIASES)

    if row_matches(probe.columns):
        return None                                    # already correct

    for i in range(min(scan, len(probe))):
        if row_matches(probe.iloc[i].tolist()):
            return i + 1                               # header is the row after the skip
    return None


def _pick_columns(df: pd.DataFrame, date_column: Optional[str], price_column: Optional[str],
                  findings: list[Finding]) -> tuple[str, str]:
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = ["_".join(str(x) for x in c if str(x) != "nan").strip("_")
                      for c in df.columns]
    lookup = {_norm_header(c): c for c in df.columns}

    dcol = date_column if date_column in df.columns else None
    pcol = price_column if price_column in df.columns else None
    if date_column and dcol is None:
        findings.append(_warn(FindingCode.COLUMNS_ASSUMED,
                              f"Requested date column '{date_column}' is not present; "
                              "falling back to auto-detection.",
                              available=[str(c) for c in df.columns]))
    if price_column and pcol is None:
        findings.append(_warn(FindingCode.COLUMNS_ASSUMED,
                              f"Requested price column '{price_column}' is not present; "
                              "falling back to auto-detection.",
                              available=[str(c) for c in df.columns]))

    if dcol is None:
        dcol = next((lookup[k] for k in lookup if k in _DATE_ALIASES), None)
    if pcol is None:
        pcol = next((lookup[k] for k in lookup
                     if k in _PRICE_ALIASES and lookup[k] != dcol), None)

    if dcol is None or pcol is None:
        if df.shape[1] < 2:
            raise SRRIInputError(_err(
                FindingCode.NO_DATE_COLUMN,
                "The file has fewer than two columns; a date column and a price column are required.",
                "Provide a file with at least a date column and a NAV/price column.",
                columns=[str(c) for c in df.columns]))
        dcol = dcol or df.columns[0]
        pcol = pcol if pcol is not None and pcol != dcol else df.columns[1]
        findings.append(_info(
            FindingCode.COLUMNS_ASSUMED,
            f"Column names were not recognised — assuming '{dcol}' is the date "
            f"and '{pcol}' is the price.",
            date_column=str(dcol), price_column=str(pcol),
            available=[str(c) for c in df.columns]))
    return dcol, pcol


def _inspect_date_tokens(raw_values: pd.Series) -> tuple[Optional[DateFormat], dict]:
    """Look at the literal text of the date column.

    Returns (implied_format, stats).  `implied_format` is DMY / MDY when the
    data itself settles the question (a component > 12 appears), ISO when the
    strings are ISO-shaped, and None when genuinely ambiguous.
    """
    stats = {"first_gt_12": 0, "second_gt_12": 0, "iso": 0, "parsed_tokens": 0}
    if pd.api.types.is_datetime64_any_dtype(raw_values):
        return DateFormat.ISO, stats | {"native_datetime": True}

    for v in raw_values.dropna().astype(str).head(5000):
        m_iso = _ISO_LIKE.match(v)
        if m_iso:
            stats["iso"] += 1
            stats["parsed_tokens"] += 1
            continue
        m = _DMY_LIKE.match(v)
        if not m:
            continue
        stats["parsed_tokens"] += 1
        a, b = int(m.group(1)), int(m.group(2))
        if a > 12:
            stats["first_gt_12"] += 1
        if b > 12:
            stats["second_gt_12"] += 1

    if stats["parsed_tokens"] == 0:
        return None, stats
    if stats["iso"] and stats["iso"] == stats["parsed_tokens"]:
        return DateFormat.ISO, stats
    if stats["first_gt_12"] and stats["second_gt_12"]:
        return None, stats                      # conflicting — cannot be one format
    if stats["first_gt_12"]:
        return DateFormat.DMY, stats
    if stats["second_gt_12"]:
        return DateFormat.MDY, stats
    return None, stats                          # every component <= 12 -> ambiguous


def _parse_dates(raw_values: pd.Series, requested: DateFormat,
                 findings: list[Finding]) -> tuple[pd.Series, DateFormat]:
    implied, stats = _inspect_date_tokens(raw_values)
    resolved = requested

    if requested is DateFormat.AUTO:
        if implied is None:
            code = (FindingCode.DATE_FORMAT_CONFLICT
                    if stats["first_gt_12"] and stats["second_gt_12"]
                    else FindingCode.DATE_FORMAT_AMBIGUOUS)
            findings.append(_err(
                code,
                "The date format cannot be inferred from this file "
                "(for example '03/04/2024' is both 3 April and 4 March).",
                "Re-submit stating explicitly whether dates are DD/MM/YYYY or MM/DD/YYYY.",
                **stats))
            resolved = DateFormat.DMY          # parse anyway so other checks can run
        else:
            resolved = implied
            findings.append(_info(FindingCode.DATE_FORMAT_INFERRED,
                                  f"Date format inferred from the data: {implied.value.upper()}.",
                                  **stats))
    else:
        if implied is not None and implied is not requested and implied is not DateFormat.ISO:
            findings.append(_err(
                FindingCode.DATE_FORMAT_CONFLICT,
                f"You selected {requested.value.upper()} but the file contains dates that "
                f"are only valid as {implied.value.upper()} "
                f"(a day/month component greater than 12).",
                "Re-submit with the correct date format — parsing this file as "
                f"{requested.value.upper()} would produce a wrong SRRI.",
                requested=requested.value, implied=implied.value, **stats))
        elif implied is None and stats["parsed_tokens"] > 0 and requested is not DateFormat.ISO:
            findings.append(_warn(
                FindingCode.DATE_FORMAT_ASSUMED,
                f"Every date in this file is valid under both DD/MM and MM/DD; it has been "
                f"parsed as {requested.value.upper()} as instructed. This cannot be verified "
                "from the data.",
                "Confirm with the data provider which convention the file uses.",
                requested=requested.value, **stats))

    if resolved is DateFormat.ISO:
        parsed = pd.to_datetime(raw_values, errors="coerce", yearfirst=True, dayfirst=False)
    else:
        parsed = pd.to_datetime(raw_values, errors="coerce",
                                dayfirst=(resolved is DateFormat.DMY))
    return parsed, resolved


def read_prices(source: Source,
                *,
                date_format: DateFormat = DateFormat.DMY,
                filename: Optional[str] = None,
                sheet: Optional[Union[str, int]] = None,
                skiprows: int = 0,
                date_column: Optional[str] = None,
                price_column: Optional[str] = None) -> PriceSeries:
    """Parse an uploaded NAV file into a clean price series.

    Accepts raw bytes, a file object, a filesystem path, or a pandas object —
    a web server never has the file on disk the way a CLI does.

    This function *parses*.  It does not judge the data: everything it notices
    becomes a `Finding`, and the regulatory sufficiency checks live in
    `validate()`.
    """
    findings: list[Finding] = []
    raw, sha, kind, fname = _materialise(source, filename)

    if not raw:
        raise SRRIInputError(_err(FindingCode.NO_DATA, "The uploaded file is empty."))

    if isinstance(source, (pd.DataFrame, pd.Series)):
        df = source.to_frame() if isinstance(source, pd.Series) else source.copy()
        if isinstance(df.index, pd.DatetimeIndex):
            df = df.reset_index()
    else:
        kind = _sniff_kind(raw, fname)
        if skiprows == 0:
            detected = _find_header_row(raw, kind, fname, sheet)
            if detected:
                findings.append(_info(
                    FindingCode.COLUMNS_ASSUMED,
                    f"A metadata block was detected above the data — reading the "
                    f"column headers from row {detected + 1} and the first "
                    f"observation from row {detected + 2}.",
                    skiprows_applied=detected))
                skiprows = detected
        df = _read_tabular(raw, kind, fname, sheet, skiprows)

    raw_rows = len(df)
    if raw_rows == 0:
        raise SRRIInputError(_err(FindingCode.NO_DATA,
                                  "The file contains no rows below the header.",
                                  "Check the 'rows to skip' setting.", skiprows=skiprows))

    dcol, pcol = _pick_columns(df, date_column, price_column, findings)

    work = df[[dcol, pcol]].copy()
    work.columns = ["Date", "Price"]

    parsed_dates, resolved_fmt = _parse_dates(work["Date"], date_format, findings)
    work["Date"] = parsed_dates
    work["Price"] = pd.to_numeric(work["Price"], errors="coerce")

    n_bad_dates = int(work["Date"].isna().sum())
    n_bad_prices = int(work["Price"].isna().sum() - (work["Date"].isna() & work["Price"].isna()).sum())

    if n_bad_dates == raw_rows:
        raise SRRIInputError(_err(
            FindingCode.ALL_DATES_UNPARSEABLE,
            f"None of the {raw_rows:,} values in column '{dcol}' could be read as dates.",
            "Check that the correct column and date format were selected.",
            column=str(dcol), date_format=resolved_fmt.value))
    if int(work["Price"].isna().sum()) == raw_rows:
        raise SRRIInputError(_err(
            FindingCode.ALL_PRICES_UNPARSEABLE,
            f"None of the {raw_rows:,} values in column '{pcol}' could be read as numbers.",
            "Check that the correct price/NAV column was selected.", column=str(pcol)))

    if n_bad_dates:
        findings.append(_warn(
            FindingCode.DATES_DROPPED,
            f"{n_bad_dates:,} of {raw_rows:,} rows had an unreadable date and were dropped.",
            "Review the source file for blank rows, footers, or text in the date column.",
            dropped=n_bad_dates, total=raw_rows, column=str(dcol)))
    if n_bad_prices > 0:
        findings.append(_warn(
            FindingCode.PRICES_DROPPED,
            f"{n_bad_prices:,} rows had an unreadable price and were dropped.",
            "Review the source file for 'N/A', blanks, or text in the price column.",
            dropped=n_bad_prices, total=raw_rows, column=str(pcol)))

    clean = work.dropna(subset=["Date", "Price"])

    non_positive = int((clean["Price"] <= 0).sum())
    if non_positive:
        findings.append(_warn(
            FindingCode.NON_POSITIVE_PRICES,
            f"{non_positive:,} rows had a zero or negative price and were dropped.",
            "A NAV of zero or below usually indicates a data error at source.",
            dropped=non_positive))
        clean = clean[clean["Price"] > 0]

    if not clean["Date"].is_monotonic_increasing:
        findings.append(_info(FindingCode.UNSORTED_DATES,
                              "Rows were not in date order and have been sorted."))
    clean = clean.sort_values("Date")

    dupes = int(clean["Date"].duplicated().sum())
    if dupes:
        findings.append(_warn(
            FindingCode.DUPLICATE_DATES,
            f"{dupes:,} duplicate dates were found; the last value for each date was kept.",
            "Confirm with the data provider which observation is authoritative.",
            duplicates=dupes))
    clean = clean.drop_duplicates("Date", keep="last")

    prices = pd.Series(clean["Price"].to_numpy(dtype=float),
                       index=pd.DatetimeIndex(clean["Date"], name="Date"),
                       name="Price")

    if prices.empty:
        raise SRRIInputError(_err(
            FindingCode.NO_DATA,
            "No usable observations remain after cleaning.",
            "Check the column selection, the date format, and the file contents."))

    return PriceSeries(
        prices=prices, findings=findings, sha256=sha, filename=fname,
        n_bytes=len(raw), input_kind=kind, raw_rows=raw_rows,
        date_format_requested=date_format, date_format_resolved=resolved_fmt,
        date_column=str(dcol), price_column=str(pcol),
    )


# ======================================================================
# Step 2 — validate (its own step, returns tagged findings)
# ======================================================================

def detect_cadence(index: pd.DatetimeIndex) -> tuple[InputCadence, float]:
    gaps = pd.Series(index).diff().dt.days.dropna()
    med = float(gaps.median()) if len(gaps) else float("nan")
    if np.isnan(med):
        return InputCadence.IRREGULAR, med
    if med <= DAILY_MAX_GAP:
        return InputCadence.DAILY, med
    if WEEKLY_MIN_GAP <= med <= WEEKLY_MAX_GAP:
        return InputCadence.WEEKLY, med
    if MONTHLY_MIN_GAP <= med <= MONTHLY_MAX_GAP:
        return InputCadence.MONTHLY, med
    return InputCadence.IRREGULAR, med


def resolve_frequency(requested: Frequency, cadence: InputCadence,
                      findings: list[Finding]) -> Frequency:
    """AUTO prefers weekly, per CESR Box 1 §2 ('use weekly returns; only use
    monthly where weekly NAV is unavailable')."""
    if requested is not Frequency.AUTO:
        return requested
    chosen = Frequency.MONTHLY if cadence is InputCadence.MONTHLY else Frequency.WEEKLY
    findings.append(_info(
        FindingCode.FREQUENCY_AUTO_SELECTED,
        f"Input cadence is {cadence.value}; calculating on the "
        f"{chosen.value} basis (CESR Box 1 §2 prefers weekly where available).",
        cadence=cadence.value, chosen=chosen.value))
    return chosen


def validate(price_series: PriceSeries,
             frequency: Frequency,
             *,
             min_periods_override: Optional[MinPeriodsOverride] = None,
             max_gap_days: Optional[int] = None,
             stale_run_periods: int = 10,
             extreme_return: float = 0.50) -> list[Finding]:
    """Regulatory and data-quality validation, separate from parsing.

    Returns a flat list of findings.  ERROR findings must block the user;
    WARNING findings may be acknowledged and proceeded past.
    """
    findings: list[Finding] = []
    prices = price_series.prices
    n = len(prices)

    cadence, med_gap = detect_cadence(prices.index)
    findings.append(_info(
        FindingCode.CADENCE_DETECTED,
        f"Input cadence detected as {cadence.value} "
        f"(median gap {med_gap:.1f} days across {n:,} observations).",
        cadence=cadence.value, median_gap_days=med_gap, observations=n))

    freq = resolve_frequency(frequency, cadence, findings)
    spec = FREQUENCY_SPEC[freq]

    # --- cadence vs chosen basis -------------------------------------
    if cadence is InputCadence.IRREGULAR:
        findings.append(_err(
            FindingCode.CADENCE_IRREGULAR,
            f"The observation dates are irregular (median gap {med_gap:.1f} days), so the "
            "series cannot be placed on a regulatory period grid with confidence.",
            "Supply a consistently dated daily, weekly, or monthly NAV series.",
            median_gap_days=med_gap))
    elif freq is Frequency.WEEKLY and cadence is InputCadence.MONTHLY:
        findings.append(_err(
            FindingCode.CADENCE_TOO_COARSE,
            "A weekly SRRI (m=52, T=260) cannot be produced from monthly data.",
            "Either supply daily/weekly NAV, or calculate on the monthly basis (m=12, T=60).",
            cadence=cadence.value, requested_basis=freq.value))
    elif freq is Frequency.MONTHLY and cadence in (InputCadence.DAILY, InputCadence.WEEKLY):
        findings.append(_warn(
            FindingCode.MONTHLY_BASIS_NOT_PREFERRED,
            f"This file contains {cadence.value} NAV, but the monthly basis (m=12, T=60) was "
            "selected. CESR Box 1 §2 requires weekly returns and permits monthly only where "
            "weekly NAV is unavailable.",
            "Calculate on the weekly basis unless there is a documented reason not to.",
            cadence=cadence.value))

    # --- history sufficiency -----------------------------------------
    span_years = (prices.index[-1] - prices.index[0]).days / 365.25
    min_periods = spec.window
    if min_periods_override is not None:
        min_periods = min_periods_override.min_periods

    # observations that will exist on the calculation grid
    grid = prices.resample(spec.resample_rule).last().dropna()
    usable_returns = max(len(grid) - 1, 0)

    if usable_returns < min_periods:
        findings.append(_err(
            FindingCode.INSUFFICIENT_HISTORY,
            f"Only {usable_returns:,} {spec.period_noun_plural} of returns are available "
            f"({span_years:.1f} years); CESR Box 1 §4 requires T={spec.window} "
            f"({spec.window // spec.m} years) for a compliant SRRI.",
            "Extend the history, or obtain an approved minimum-period override.",
            available_periods=usable_returns, required_periods=spec.window,
            effective_minimum=min_periods, history_years=round(span_years, 2)))
    elif usable_returns < spec.window:
        # only reachable with a valid override
        findings.append(_warn(
            FindingCode.SHORT_HISTORY,
            f"The SRRI is being produced on {usable_returns:,} {spec.period_noun_plural} "
            f"({span_years:.1f} years), below the CESR minimum of T={spec.window}.",
            "This figure is non-standard and must be disclosed as such.",
            available_periods=usable_returns, required_periods=spec.window,
            history_years=round(span_years, 2)))

    if min_periods_override is not None:
        findings.append(_warn(
            FindingCode.MIN_PERIODS_OVERRIDE_APPLIED,
            f"The regulatory {spec.window}-{spec.period_noun} window has been relaxed to "
            f"{min_periods_override.min_periods} by {min_periods_override.approved_by}: "
            f"{min_periods_override.reason}",
            "This override is recorded in the result's audit block and on the workbook's "
            "Audit & Findings sheet.",
            **min_periods_override.model_dump(mode="json")))

    if span_years < 5:
        findings.append(_warn(
            FindingCode.SHORT_HISTORY,
            f"The file spans {span_years:.1f} years. CESR requires a 5-year history for a "
            "full SRRI.",
            "Supplement with an appropriate benchmark or proxy series where permitted.",
            history_years=round(span_years, 2)))

    # --- gaps, staleness, outliers -----------------------------------
    gaps = pd.Series(prices.index).diff().dt.days.dropna()
    limit = max_gap_days if max_gap_days is not None else (
        45 if cadence is InputCadence.MONTHLY else 21)
    big = gaps[gaps > limit]
    if len(big):
        worst = int(big.max())
        at = prices.index[int(big.idxmax())].date()
        findings.append(_warn(
            FindingCode.LARGE_GAP,
            f"{len(big):,} gap(s) longer than {limit} days in the price history "
            f"(largest {worst} days, ending {at}).",
            "Missing NAV points distort the volatility window; confirm the history is complete.",
            gap_count=int(len(big)), largest_gap_days=worst, gap_end=str(at),
            threshold_days=limit))

    unchanged = (prices.diff() == 0)
    if unchanged.any():
        runs = unchanged.groupby((~unchanged).cumsum()).cumsum()
        longest = int(runs.max())
        if longest >= stale_run_periods:
            # positional, not label-based: the index is a DatetimeIndex
            at = prices.index[int(np.argmax(runs.to_numpy()))].date()
            findings.append(_warn(
                FindingCode.STALE_PRICES,
                f"The NAV is unchanged for {longest} consecutive observations "
                f"(ending {at}), which understates measured volatility.",
                "Check for carried-forward or stale pricing at source.",
                longest_run=longest, run_end=str(at)))

    rets = prices.pct_change().dropna()
    extreme = rets[rets.abs() > extreme_return]
    if len(extreme):
        worst_idx = extreme.abs().idxmax()
        findings.append(_warn(
            FindingCode.EXTREME_RETURN,
            f"{len(extreme):,} single-period move(s) larger than "
            f"{extreme_return:.0%} (largest {extreme.loc[worst_idx]:+.1%} on "
            f"{worst_idx.date()}).",
            "Verify these are genuine market moves and not unadjusted distributions "
            "or share-class events.",
            count=int(len(extreme)), largest_pct=round(float(extreme.loc[worst_idx]) * 100, 2),
            date=str(worst_idx.date())))

    if n < 10:
        findings.append(_err(
            FindingCode.TOO_FEW_OBSERVATIONS,
            f"Only {n} usable observations — too few to calculate anything meaningful.",
            "Supply a complete NAV history.", observations=n))

    return findings


# ======================================================================
# Step 3 — calculate (returns a result object; never touches the disk)
# ======================================================================

def _build_grid(prices: pd.Series, spec: FrequencySpec,
                cadence: InputCadence, findings: list[Finding]) -> pd.DataFrame:
    """Place the price series on the regulatory period grid."""
    already = ((spec.frequency is Frequency.WEEKLY and cadence is InputCadence.WEEKLY)
               or (spec.frequency is Frequency.MONTHLY and cadence is InputCadence.MONTHLY))
    grid = prices.resample(spec.resample_rule).last().dropna().to_frame("Price")
    if not already:
        findings.append(_info(
            FindingCode.RESAMPLED,
            f"{cadence.value.title()} input resampled to {len(grid):,} "
            f"{spec.period_noun} points ({spec.basis_label}).",
            rule=spec.resample_rule, points=len(grid)))
    # Simple arithmetic returns: r = (P_t - P_{t-1}) / P_{t-1}
    # CESR Box 1 example: (96-100)/100 = -4.00% -> simple, not log.
    grid["Return"] = grid["Price"].pct_change()
    return grid


def calculate(price_series: PriceSeries,
              frequency: Frequency,
              *,
              findings: Optional[list[Finding]] = None,
              min_periods_override: Optional[MinPeriodsOverride] = None,
              buffer_months: int = BUFFER_MONTHS_DEFAULT,
              sheet: Optional[Union[str, int]] = None,
              skiprows: int = 0,
              verify_parity: bool = True) -> SRRIResult:
    """Compute the SRRI series and return it.  No file is written.

    `findings` should be the combined output of `read_prices` and `validate`;
    if any of them is an ERROR the result comes back with status BLOCKED and an
    empty series, so the caller can render the blocking reasons without
    catching an exception.
    """
    all_findings: list[Finding] = list(findings or [])
    prices = price_series.prices
    cadence, _ = detect_cadence(prices.index)
    freq = resolve_frequency(frequency, cadence, all_findings)
    spec = FREQUENCY_SPEC[freq]

    min_periods = (min_periods_override.min_periods
                   if min_periods_override is not None else spec.window)

    audit = AuditInfo(
        input_sha256=price_series.sha256,
        input_filename=price_series.filename,
        input_bytes=price_series.n_bytes,
        input_kind=price_series.input_kind,
        frequency=freq, m=spec.m, window=spec.window,
        annualisation=spec.annualisation,
        date_format=price_series.date_format_requested,
        date_format_resolved=price_series.date_format_resolved,
        buffer_months=buffer_months,
        min_periods=min_periods,
        min_periods_is_regulatory_default=(min_periods_override is None),
        min_periods_override=min_periods_override,
        sheet=sheet, skiprows=skiprows,
    )

    common = dict(
        audit=audit, findings=all_findings,
        input_cadence=cadence,
        input_first_date=prices.index[0].date(),
        input_last_date=prices.index[-1].date(),
        input_rows=price_series.raw_rows,
        history_years=round((prices.index[-1] - prices.index[0]).days / 365.25, 2),
    )

    if any(f.severity is Severity.ERROR for f in all_findings):
        return SRRIResult(status=ResultStatus.BLOCKED, **common)

    grid = _build_grid(prices, spec, cadence, all_findings)
    rets = grid["Return"]

    # Vectorised equivalent of the Box 1 §4 formula:
    #   sqrt( (m/(T-1)) * SUM (r - r_bar)^2 )  ==  std(r, ddof=1) * sqrt(m)
    # ~150x faster than the per-point Python loop the old scripts used on a
    # 20-year weekly history (72 ms -> 0.5 ms), and the gap widens with length
    # because the loop is O(n*T) and this is O(n).  Parity with the literal
    # formula is asserted below rather than assumed.
    roll = rets.rolling(window=spec.window, min_periods=min_periods)
    ann_vol = roll.std(ddof=1) * np.sqrt(spec.m)
    mean_ret = roll.mean()
    n_in_window = rets.rolling(window=spec.window, min_periods=1).count().fillna(0)

    if verify_parity:
        _assert_parity(rets, ann_vol, spec, min_periods)

    srri_raw = ann_vol.map(classify_srri)
    srri_raw = pd.Series(
        [np.nan if v is None else float(v) for v in srri_raw],
        index=ann_vol.index)

    disclosed = apply_buffer_zone(srri_raw, buffer_months)

    run_id = (srri_raw != srri_raw.shift()).cumsum()
    periods_at = srri_raw.groupby(run_id).cumcount() + 1

    points: list[SeriesPoint] = []
    for ts in grid.index:
        raw_v = srri_raw.loc[ts]
        dis_v = disclosed.loc[ts]
        valid = pd.notna(raw_v)
        points.append(SeriesPoint(
            date=ts.date(),
            price=float(grid.at[ts, "Price"]),
            period_return_pct=_pct(rets.loc[ts]),
            periods_in_window=int(n_in_window.loc[ts]),
            mean_return_pct=_pct(mean_ret.loc[ts]),
            period_vol_pct=_pct(ann_vol.loc[ts] / np.sqrt(spec.m)) if valid else None,
            ann_vol_pct=_pct(ann_vol.loc[ts]),
            srri_raw=int(raw_v) if valid else None,
            srri_disclosed=int(dis_v) if pd.notna(dis_v) else None,
            risk_description=(RISK_LABELS.get(int(dis_v), "N/A")
                              if pd.notna(dis_v) else "Insufficient Data"),
            periods_at_srri=int(periods_at.loc[ts]) if valid else 0,
            status="Valid" if valid else "Insufficient Data",
        ))

    valid_points = [p for p in points if p.status == "Valid"]
    distribution = {c: sum(1 for p in valid_points if p.srri_disclosed == c)
                    for c in range(1, 8)}

    if not valid_points:
        return SRRIResult(status=ResultStatus.NO_VALID_SRRI,
                          n_periods=len(points), n_valid_periods=0,
                          series=points, distribution=distribution, **common)

    last = valid_points[-1]
    status = (ResultStatus.OK_WITH_WARNINGS
              if any(f.severity is Severity.WARNING for f in all_findings)
              else ResultStatus.OK)

    return SRRIResult(
        status=status,
        as_of_date=last.date,
        annualised_volatility=(last.ann_vol_pct / 100 if last.ann_vol_pct is not None else None),
        srri_raw=last.srri_raw,
        srri_disclosed=last.srri_disclosed,
        risk_description=last.risk_description,
        n_periods=len(points),
        n_valid_periods=len(valid_points),
        first_valid_date=valid_points[0].date,
        series=points, distribution=distribution,
        **common,
    )


def _pct(v) -> Optional[float]:
    return None if v is None or pd.isna(v) else float(v) * 100


def _assert_parity(rets: pd.Series, ann_vol: pd.Series, spec: FrequencySpec,
                   min_periods: int, sample: int = 8) -> None:
    """Cross-check the vectorised path against the literal Box 1 §4 formula.

    Cheap insurance: if a pandas upgrade ever changes rolling-window semantics,
    this fails loudly rather than quietly shifting a published SRRI.
    """
    valid_idx = ann_vol.dropna().index
    if len(valid_idx) == 0:
        return
    step = max(len(valid_idx) // sample, 1)
    for ts in valid_idx[::step]:
        pos = rets.index.get_loc(ts)
        window = rets.iloc[max(0, pos - spec.window + 1): pos + 1].dropna()
        if len(window) < min_periods:
            continue
        reference = cesr_volatility(window.values, spec.m)
        if not np.isclose(reference, ann_vol.loc[ts], rtol=1e-12, atol=1e-15):
            raise AssertionError(
                f"CESR Box 1 §4 parity check failed at {ts.date()}: "
                f"literal formula {reference!r} vs vectorised {ann_vol.loc[ts]!r}")


# ======================================================================
# One-call convenience wrapper
# ======================================================================

def run(source: Source,
        frequency: Frequency = Frequency.AUTO,
        *,
        date_format: DateFormat = DateFormat.DMY,
        filename: Optional[str] = None,
        sheet: Optional[Union[str, int]] = None,
        skiprows: int = 0,
        date_column: Optional[str] = None,
        price_column: Optional[str] = None,
        min_periods_override: Optional[MinPeriodsOverride] = None,
        buffer_months: int = BUFFER_MONTHS_DEFAULT,
        verify_parity: bool = True) -> SRRIResult:
    """read -> validate -> calculate, in one call.

    Never raises for recoverable input problems; check `result.is_blocked` and
    render `result.errors`.  Only a genuinely unreadable input raises
    `SRRIInputError`.
    """
    ps = read_prices(source, date_format=date_format, filename=filename,
                     sheet=sheet, skiprows=skiprows,
                     date_column=date_column, price_column=price_column)
    checks = validate(ps, frequency, min_periods_override=min_periods_override)
    return calculate(ps, frequency,
                     findings=list(ps.findings) + checks,
                     min_periods_override=min_periods_override,
                     buffer_months=buffer_months,
                     sheet=sheet, skiprows=skiprows,
                     verify_parity=verify_parity)


# ======================================================================
# Optional Excel export — an artifact derived from a result, not the result
# ======================================================================

SRRI_HEX = {
    1: "4CAF50", 2: "8BC34A", 3: "CDDC39",
    4: "FFC107", 5: "FF9800", 6: "FF5722", 7: "F44336",
}
SEVERITY_HEX = {
    Severity.ERROR: "F44336",
    Severity.WARNING: "FFC107",
    Severity.INFO: "E8F0F5",
}


def export_workbook(result: SRRIResult,
                    destination: Optional[Union[str, Path, IO[bytes]]] = None) -> bytes:
    """Build the audit workbook from a result and return it as bytes.

    Sheets: Summary | SRRI Calculations | Distribution | Methodology |
            Audit & Findings.

    The first four are unchanged in substance from the original scripts.
    "Audit & Findings" is new: it carries the engine version, the input hash,
    the resolved parameters, any override attribution, and every finding —
    so the workbook attached to a published document is self-describing.

    `destination` is optional; bytes are always returned, which is what a
    web response needs.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    grey = PatternFill("solid", start_color="F0F0F0", end_color="F0F0F0")
    blue = PatternFill("solid", start_color="E8F0F5", end_color="E8F0F5")

    spec = FREQUENCY_SPEC[result.audit.frequency]
    noun = spec.period_noun.title()
    nouns = spec.period_noun_plural.title()

    def header(ws, row, labels):
        for c, label in enumerate(labels, 1):
            cell = ws.cell(row=row, column=c, value=label)
            cell.fill, cell.font, cell.border = hdr_fill, hdr_font, border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def srri_colour(cell, v):
        if v is not None and pd.notna(v):
            iv = int(v)
            cell.fill = PatternFill("solid", start_color=SRRI_HEX[iv], end_color=SRRI_HEX[iv])
            if iv >= 6:
                cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)

    def key_value_sheet(ws, rows, widths=(42, 78)):
        for off, (label, value, is_section) in enumerate(rows, 3):
            a = ws.cell(row=off, column=1, value=label)
            b = ws.cell(row=off, column=2, value=value)
            a.font = Font(name="Arial", size=10)
            b.font = Font(name="Arial", size=10)
            b.alignment = Alignment(wrap_text=True, vertical="top")
            if is_section and label:
                a.font = Font(bold=True, size=11, name="Arial")
                a.fill = blue
                b.fill = blue
        ws.column_dimensions["A"].width = widths[0]
        ws.column_dimensions["B"].width = widths[1]

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ---------------- Summary ----------------
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"SRRI {spec.frequency.value.upper()} CALCULATION SUMMARY"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E78", name="Arial")

    a = result.audit
    valid = [p for p in result.series if p.status == "Valid"]
    n_valid = len(valid)
    entries = [
        ("METHOD (CESR/10-673)", "", True),
        ("Basis", spec.basis_label, False),
        ("Input frequency", result.input_cadence.value.upper() if result.input_cadence else "N/A", False),
        ("Periods per year (m)", str(a.m), False),
        (f"Window (T)", f"{a.window} {spec.period_noun_plural} (5 years)", False),
        ("Annualisation", f"sigma_{spec.period_noun} x sqrt({a.m})", False),
        ("Min window applied", f"{a.min_periods} {spec.period_noun_plural}"
         + ("" if a.min_periods_is_regulatory_default else "  (OVERRIDE — see Audit sheet)"), False),
        ("Buffer (Box 3)", f"{a.buffer_months} months", False),
        ("", "", False),
        ("DATASET", "", True),
        ("Input range", f"{result.input_first_date} to {result.input_last_date}", False),
        ("History", f"{result.history_years} years", False),
        (f"Total {spec.period_noun_plural}", f"{result.n_periods:,}", False),
        (f"Valid SRRI {spec.period_noun_plural}", f"{n_valid:,}", False),
        ("First valid SRRI", str(result.first_valid_date) if result.first_valid_date else "N/A", False),
        ("", "", False),
        ("LATEST RESULT", "", True),
        (f"{noun} ending", str(result.as_of_date) if result.as_of_date else "N/A", False),
        ("NAV Price", f"{valid[-1].price:.5f}" if valid else "N/A", False),
        ("Annualised Volatility",
         f"{result.annualised_volatility_pct:.2f}%" if result.annualised_volatility_pct is not None else "N/A", False),
        ("SRRI — Raw", str(result.srri_raw) if result.srri_raw is not None else "N/A", False),
        ("SRRI — Disclosed (Box 3)",
         str(result.srri_disclosed) if result.srri_disclosed is not None else "N/A", False),
        ("Risk Description", result.risk_description or "N/A", False),
        ("Result status", result.status.value.upper(), False),
        ("", "", False),
        ("SRRI DISTRIBUTION (Disclosed)", "", True),
    ]
    for cls in range(1, 8):
        cnt = result.distribution.get(cls, 0)
        pctg = cnt / n_valid * 100 if n_valid else 0.0
        entries.append((f"  SRRI {cls}  {RISK_LABELS[cls]}", f"{cnt:,}  ({pctg:.1f}%)", False))
    entries += [
        ("", "", False),
        ("PROVENANCE", "", True),
        ("Engine", f"{a.engine_name} v{a.engine_version}", False),
        ("Input SHA-256", a.input_sha256, False),
        ("Calculated at (UTC)", a.calculated_at.strftime("%Y-%m-%d %H:%M:%S"), False),
    ]
    key_value_sheet(ws, entries, widths=(42, 70))
    for off, (label, _v, _s) in enumerate(entries, 3):
        if label == "SRRI — Disclosed (Box 3)" and result.srri_disclosed is not None:
            d = int(result.srri_disclosed)
            cell = ws.cell(row=off, column=2)
            cell.font = Font(bold=True, size=13, color="FFFFFF", name="Arial")
            cell.fill = PatternFill("solid", start_color=SRRI_HEX[d], end_color=SRRI_HEX[d])

    # ---------------- SRRI Calculations ----------------
    ws = wb.create_sheet("SRRI Calculations")
    ws["A1"] = (f"SRRI {spec.frequency.value.upper()} CALCULATIONS  |  CESR/10-673 Box 1 §4  |  "
                f"sigma = sqrt[ ({a.m}/(T-1)) x SUM(r - r_bar)^2 ]  |  "
                f"Window: {a.window} {spec.period_noun_plural}")
    ws["A1"].font = Font(bold=True, size=11, color="1F4E78", name="Arial")
    H = 3
    header(ws, H, [
        f"{noun} Ending", "Price (NAV)", f"{noun} Return (%)",
        f"{nouns} in Window", "Mean Return (%)", f"{noun} Vol (%)",
        "Ann. Vol (%)", "SRRI (Raw)", "SRRI (Disclosed)",
        "Risk Description", f"{nouns} at SRRI", "Status",
    ])
    ws.row_dimensions[H].height = 32
    fmts = ["yyyy-mm-dd", "0.00000", "0.0000", "0", "0.0000",
            "0.0000", "0.00", "0", "0", "@", "0", "@"]
    for r, p in enumerate(result.series, H + 1):
        vals = [p.date, p.price, p.period_return_pct, p.periods_in_window,
                p.mean_return_pct, p.period_vol_pct, p.ann_vol_pct,
                p.srri_raw, p.srri_disclosed, p.risk_description,
                p.periods_at_srri or None, p.status]
        for c, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.number_format = fmt
            cell.border = border
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(name="Arial", size=9)
            if p.status == "Insufficient Data":
                cell.fill = grey
            if c == 8:
                srri_colour(cell, p.srri_raw)
            if c == 9:
                srri_colour(cell, p.srri_disclosed)
    for c, w in enumerate([12, 12, 16, 15, 14, 14, 12, 11, 14, 20, 14, 16], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = f"A{H + 1}"

    # ---------------- Distribution ----------------
    ws = wb.create_sheet("Distribution")
    ws["A1"] = (f"SRRI DISTRIBUTION — {spec.frequency.value.title()} "
                "(Disclosed after Box 3 buffer)")
    ws["A1"].font = Font(bold=True, size=11, color="1F4E78", name="Arial")
    header(ws, 3, ["SRRI", "Description", "Volatility Range", nouns, "% of Time"])
    for off, cls in enumerate(range(1, 8), 4):
        lo, hi = next((lo, hi) for s, lo, hi in SRRI_BANDS if s == cls)
        cnt = result.distribution.get(cls, 0)
        share = cnt / n_valid if n_valid else 0.0
        hi_str = f"< {hi * 100:.2f}%" if hi != float("inf") else ">= 25.00%"
        for c, v in enumerate([f"SRRI {cls}", RISK_LABELS[cls],
                               f">= {lo * 100:.2f}%  {hi_str}", cnt, share], 1):
            cell = ws.cell(row=off, column=c, value=v)
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            if c == 5:
                cell.number_format = "0.00%"
            if c == 1:
                srri_colour(cell, cls)
    for c, w in zip("ABCDE", [12, 22, 24, 12, 12]):
        ws.column_dimensions[c].width = w

    # ---------------- Methodology ----------------
    ws = wb.create_sheet("Methodology")
    ws["A1"] = (f"{spec.frequency.value.upper()} SRRI METHODOLOGY — "
                "CESR/10-673 (ESMA, 1 July 2010)")
    ws["A1"].font = Font(bold=True, size=12, color="1F4E78", name="Arial")
    rows = [
        ("REGULATORY BASIS", "", True),
        ("Document", "CESR/10-673, 1 July 2010", False),
        ("Implementing reg.", "Commission Regulation (EU) No 583/2010", False),
        ("Sections used", "Box 1 §4 (formula), Box 2 (thresholds), Box 3 (buffer)", False),
        ("", "", False),
        ("BOX 1 §4 — THE EXACT FORMULA", "", True),
        ("Volatility", "sigma = sqrt[ (m/(T-1)) x SUM(r_t - r_bar)^2 ]", False),
        (f"m ({spec.frequency.value})", f"{a.m}  — periods per year", False),
        (f"T ({spec.frequency.value})", f"{a.window} — {spec.period_noun_plural} in the 5-year window", False),
        ("Regulation quote", f'"m={a.m} and T={a.window} for {spec.period_noun} returns"', False),
        ("Implementation", f"rolling std(returns, ddof=1) x sqrt({a.m}); cross-checked "
                           "against the literal formula on a sample of points at run time", False),
        ("", "", False),
        ("INPUT HANDLING", "", True),
        ("Detected cadence", result.input_cadence.value if result.input_cadence else "N/A", False),
        ("Grid construction", f"resample('{spec.resample_rule}').last() — {spec.basis_label}", False),
        ("Date format", f"requested {a.date_format.value.upper()}, "
                        f"applied {a.date_format_resolved.value.upper()}", False),
        ("", "", False),
        ("RETURNS", "", True),
        ("Type", "Simple arithmetic: r_t = (P_t - P_{t-1}) / P_{t-1}", False),
        ("Source", "CESR Box 1 example: (96-100)/100 = -4.00% -> simple, not log", False),
        ("", "", False),
        ("BOX 2 — THRESHOLDS", "", True),
        ("SRRI 1", " 0.00% <= sigma <  0.50%  (Lowest Risk)", False),
        ("SRRI 2", " 0.50% <= sigma <  2.00%  (Very Low Risk)", False),
        ("SRRI 3", " 2.00% <= sigma <  5.00%  (Low Risk)", False),
        ("SRRI 4", " 5.00% <= sigma < 10.00%  (Medium Risk)", False),
        ("SRRI 5", "10.00% <= sigma < 15.00%  (Medium-High Risk)", False),
        ("SRRI 6", "15.00% <= sigma < 25.00%  (High Risk)", False),
        ("SRRI 7", "25.00% <= sigma           (Highest Risk)", False),
        ("", "", False),
        ("BOX 3 — BUFFER ZONE", "", True),
        ("Rule", "The disclosed SRRI changes only if the new category has held at each "
                 f"monthly reference point over the preceding {a.buffer_months} months.", False),
        ("Implementation", "Raw SRRI condensed to month-end reference points, the "
                           f"{a.buffer_months}-month rule applied, then forward-filled back onto "
                           "the calculation grid. Identical code path for both bases.", False),
        ("", "", False),
        ("BASIS SELECTION", "", True),
        ("Box 1 §2", "Use weekly returns; use monthly only where weekly NAV is unavailable.", False),
        ("This run", spec.basis_label, False),
    ]
    key_value_sheet(ws, rows, widths=(30, 88))

    # ---------------- Audit & Findings ----------------
    ws = wb.create_sheet("Audit & Findings")
    ws.sheet_view.showGridLines = False
    ws["A1"] = "AUDIT TRAIL & DATA-QUALITY FINDINGS"
    ws["A1"].font = Font(bold=True, size=12, color="1F4E78", name="Arial")

    ov = a.min_periods_override
    audit_rows = [
        ("PROVENANCE", "", True),
        ("Engine", f"{a.engine_name} v{a.engine_version}", False),
        ("Methodology reference", a.methodology_ref, False),
        ("Calculated at (UTC)", a.calculated_at.strftime("%Y-%m-%d %H:%M:%S"), False),
        ("Result status", result.status.value.upper(), False),
        ("Fingerprint", a.fingerprint, False),
        ("", "", False),
        ("INPUT", "", True),
        ("File name", a.input_filename or "(in-memory upload)", False),
        ("Input kind", a.input_kind, False),
        ("Size (bytes)", f"{a.input_bytes:,}", False),
        ("SHA-256", a.input_sha256, False),
        ("Sheet", str(a.sheet) if a.sheet is not None else "(default)", False),
        ("Header rows skipped", str(a.skiprows), False),
        ("Rows read", f"{result.input_rows:,}", False),
        ("", "", False),
        ("PARAMETERS APPLIED", "", True),
        ("Frequency", a.frequency.value, False),
        ("m / T", f"{a.m} / {a.window}", False),
        ("Buffer months", str(a.buffer_months), False),
        ("Date format requested", a.date_format.value.upper(), False),
        ("Date format applied", a.date_format_resolved.value.upper(), False),
        ("Minimum periods", str(a.min_periods), False),
        ("Regulatory default?", "YES" if a.min_periods_is_regulatory_default else "NO — OVERRIDDEN", False),
        ("", "", False),
        ("MINIMUM-PERIOD OVERRIDE", "", True),
        ("Override applied", "NO" if ov is None else "YES", False),
    ]
    if ov is not None:
        audit_rows += [
            ("Relaxed minimum", f"{ov.min_periods} {spec.period_noun_plural} "
                                f"(regulatory {a.window})", False),
            ("Approved by", ov.approved_by, False),
            ("Reason", ov.reason, False),
            ("Approved at (UTC)", ov.approved_at.strftime("%Y-%m-%d %H:%M:%S"), False),
            ("Ticket reference", ov.ticket_reference or "(none)", False),
        ]
    key_value_sheet(ws, audit_rows, widths=(30, 88))

    fr = len(audit_rows) + 5
    ws.cell(row=fr - 1, column=1, value="FINDINGS").font = Font(bold=True, size=11, name="Arial")
    header(ws, fr, ["Severity", "Code", "Message", "Suggested action"])
    ws.row_dimensions[fr].height = 24
    for i, f in enumerate(result.findings, fr + 1):
        for c, v in enumerate([f.severity.value.upper(), f.code.value, f.message,
                               f.remediation or ""], 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.border = border
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c == 1:
                cell.fill = PatternFill("solid", start_color=SEVERITY_HEX[f.severity],
                                        end_color=SEVERITY_HEX[f.severity])
                if f.severity is Severity.ERROR:
                    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    if not result.findings:
        ws.cell(row=fr + 1, column=1, value="No findings — input passed all checks.")
    for c, w in zip("ABCD", [12, 30, 80, 60]):
        ws.column_dimensions[c].width = w

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    if destination is not None:
        if hasattr(destination, "write"):
            destination.write(data)
        else:
            Path(destination).write_bytes(data)
    return data


# ======================================================================
# CLI — a thin wrapper over the library, not the entry point
# ======================================================================

def main(argv: Optional[list[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        prog="srri_engine",
        description="UCITS KIID SRRI calculator — CESR/10-673 (weekly m=52/T=260, "
                    "monthly m=12/T=60). Thin CLI over the srri_engine library.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python srri_engine.py nav_daily.xlsx  --frequency weekly  --date-format dmy\n"
            "  python srri_engine.py nav_monthly.csv --frequency monthly --date-format iso "
            "--output srri.xlsx\n"
            "  python srri_engine.py nav.csv --frequency auto --json\n"
            "\nNote: the minimum-window override is deliberately NOT a plain CLI flag.\n"
            "It requires --override-min-periods together with --override-approved-by and\n"
            "--override-reason, and the attribution is written into the result and the\n"
            "workbook. In the web app this route must be admin-only.\n"
        ),
    )
    ap.add_argument("input", help="NAV file (.xlsx/.xlsm/.xls/.csv/.txt/.tsv)")
    ap.add_argument("--frequency", choices=[f.value for f in Frequency],
                    default=Frequency.AUTO.value, help="Calculation basis (default: auto)")
    ap.add_argument("--date-format", choices=[d.value for d in DateFormat],
                    default=DateFormat.DMY.value,
                    help="Day/month ordering of the input dates (default: dmy)")
    ap.add_argument("--sheet", default=None, help="Sheet name or index (Excel only)")
    ap.add_argument("--skiprows", type=int, default=0, help="Header rows to skip")
    ap.add_argument("--date-column", default=None)
    ap.add_argument("--price-column", default=None)
    ap.add_argument("--buffer-months", type=int, default=BUFFER_MONTHS_DEFAULT)
    ap.add_argument("--output", default=None, help="Write the audit workbook here (optional)")
    ap.add_argument("--json", action="store_true", help="Emit the full result as JSON")
    ap.add_argument("--override-min-periods", type=int, default=None,
                    help="ADMIN ONLY — relax the CESR window; requires approver and reason")
    ap.add_argument("--override-approved-by", default=None)
    ap.add_argument("--override-reason", default=None)
    ap.add_argument("--override-ticket", default=None)
    ap.add_argument("-v", "--verbose", action="store_true")
    args = ap.parse_args(argv)

    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO,
                        format="%(levelname)-8s %(message)s")

    override = None
    if args.override_min_periods is not None:
        if not (args.override_approved_by and args.override_reason):
            ap.error("--override-min-periods requires --override-approved-by and "
                     "--override-reason (the relaxation must be attributable).")
        override = MinPeriodsOverride(
            min_periods=args.override_min_periods,
            approved_by=args.override_approved_by,
            reason=args.override_reason,
            ticket_reference=args.override_ticket,
        )

    sheet: Optional[Union[str, int]] = args.sheet
    if isinstance(sheet, str) and sheet.isdigit():
        sheet = int(sheet)

    try:
        result = run(Path(args.input),
                     frequency=Frequency(args.frequency),
                     date_format=DateFormat(args.date_format),
                     sheet=sheet, skiprows=args.skiprows,
                     date_column=args.date_column, price_column=args.price_column,
                     min_periods_override=override,
                     buffer_months=args.buffer_months)
    except SRRIInputError as exc:
        print(f"[ERROR] {exc.finding.code.value}: {exc.finding.message}", file=sys.stderr)
        if exc.finding.remediation:
            print(f"        {exc.finding.remediation}", file=sys.stderr)
        return 2

    if args.json:
        print(json.dumps(result.model_dump(mode="json"), indent=2, default=str))
    else:
        print(result.summary())

    if args.output:
        export_workbook(result, args.output)
        print(f"Workbook written to {args.output}")

    return 1 if result.is_blocked else 0


if __name__ == "__main__":
    sys.exit(main())
