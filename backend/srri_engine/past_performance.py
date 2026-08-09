"""
past_performance.py — UCITS KIID past-performance engine (Reg. (EU) No 583/2010)
================================================================================

Produces the calendar-year return figures behind the KIID past-performance bar
chart, from the same NAV series the SRRI engine consumes.

Regulatory basis — Commission Regulation (EU) No 583/2010, Section 4 of
Chapter III and Annex III.  The rules actually implemented here, verbatim in
substance:

  Art. 15(1)  Bar chart covering the last 10 years.
  Art. 15(2)  A UCITS with performance of less than 5 complete calendar years
              uses a presentation covering the last 5 years ONLY — not 10.
  Art. 15(3)  Years for which data is not available are shown blank, with no
              annotation other than the date.
  Art. 15(4)  With no data for one complete calendar year, no chart is shown;
              a statement of insufficient data is required instead.
  Art. 15(5)  The chart is supplemented by statements: (a) limited value as a
              guide to the future; (b) which charges are included/excluded —
              disapplied where the fund has no entry or exit charges;
              (c) the year the fund came into existence; (d) the currency.
  Art. 15(6)  No record of past performance for any part of the CURRENT
              calendar year.  Hard cut-off at 31 December of the prior year.
  Art. 16     Figures are based on the net asset value of the UCITS, on the
              basis that any distributable income has been reinvested.
  Art. 17     Where a material change occurred to objectives / investment
              policy within the displayed period, earlier performance is still
              shown, and that earlier period is labelled with a clear warning.
  Art. 18     Where the objectives section references a benchmark, a benchmark
              bar sits alongside each fund bar.  The benchmark is not shown
              for years in which the UCITS did not exist.
  Art. 19     Simulated performance is permitted only in the listed cases and
              must be prominently disclosed on the chart.
  Annex III   Y-axis linear (not logarithmic); scale adapted to the span of the
              bars and not compressed; X-axis at 0 %; every bar labelled with
              its return; figures rounded to ONE decimal place.
  Art. 23(3)  The revised past-performance presentation is published no later
              than 35 business days after 31 December each year.

House conventions (recorded because the regulation does not decide them)
-----------------------------------------------------------------------
1.  CHARGES.  Art. 16 says "based on the net asset value" and stops there.  A
    published NAV is already net of ongoing charges — management, depositary,
    administration and audit fees accrue daily as liabilities inside the fund
    before NAV per share is struck.  It is gross of entry and exit charges,
    which are levied on the investor outside the fund and never touch NAV.
    That is precisely why no charge adjustment is prescribed.  This engine
    therefore applies NONE.  It does not accept an OCF to deduct: a gross-of-
    fee track record is not a NAV series and must be rejected upstream, not
    netted down here.

2.  DISTRIBUTIONS.  Same policy as the SRRI engine (set 8 August 2026): every
    NAV series supplied is treated as accumulating or already distribution-
    adjusted.  The engine neither fetches nor infers dividends.  Both
    assumptions are written into the audit trail on every run.

3.  YEAR ANCHORING.  A calendar-year return needs the last NAV of year Y-1 and
    the last of year Y.  Weekly and monthly series rarely carry a point dated
    exactly 31 December, so the anchor is the last observation ON OR BEFORE
    31 December.  Staleness beyond a tolerance derived from the series' own
    cadence raises `STALE_YEAR_ANCHOR` — the point of which is to catch a
    truncated file, not to police normal weekly spacing.

4.  PARTIAL YEARS ARE NEVER COMPUTED.  A fund launched in May 2021 has no
    complete 2021 calendar year.  The bar is blank (Art. 15(3)); it is not a
    May-to-December stub.  Same treatment at the end of the series.

5.  CLIENT FIGURES.  Where administrator-supplied yearly figures are passed in,
    they are reconciled against the computed figures and any year differing by
    more than `RECONCILIATION_TOLERANCE_PP` raises a finding.  The engine's own
    figures populate the chart; the reconciliation is evidence, not an override.

What this module deliberately does NOT do
-----------------------------------------
- It does not splice or simulate (Art. 19).  A simulated segment must be
  supplied as part of the series by the preparer and declared via
  `simulated_through`, which drives the disclosure only.
- It does not decide whether a benchmark is required.  Art. 18 keys off the
  objectives narrative, which lives in the document, not in the NAV file.
- It does not produce structured-UCITS performance scenarios.  Art. 36(1)
  removes the past-performance section for those funds entirely.

Usage
-----
    from srri_engine import DateFormat, read_prices
    from past_performance import run, export_workbook, merge_into_workbook

    result = run("nav.xlsx",
                 reference_date=date(2026, 8, 9),
                 currency="EUR",
                 fund_inception_date=date(2013, 5, 22),
                 date_format=DateFormat.YMD)

    if result.is_blocked:
        return {"errors": [f.model_dump() for f in result.errors]}

    xlsx = export_workbook(result)                  # standalone audit workbook
    xlsx = merge_into_workbook(srri_xlsx, result)   # or appended to the SRRI one

CLI (thin wrapper, not the entry point)
---------------------------------------
    python past_performance.py nav.xlsx --currency EUR --output pp.xlsx
"""

from __future__ import annotations

import argparse
import io
import logging
import sys
from datetime import date, datetime, timezone
from enum import Enum
from pathlib import Path
from typing import IO, Any, Optional, Union

import numpy as np
import pandas as pd
from pydantic import BaseModel, ConfigDict, Field, field_validator

from srri_engine import (
    DateFormat,
    Finding,
    FindingCode,
    PriceSeries,
    Severity,
    Source,
    read_prices,
)

log = logging.getLogger("past_performance")

ENGINE_NAME = "UCITS KIID past-performance engine"
ENGINE_VERSION = "1.0.0"
METHODOLOGY_REF = (
    "Commission Regulation (EU) No 583/2010, Arts. 15-19 and Annex III"
)

# --- regulatory constants -------------------------------------------------
MAX_CHART_YEARS = 10          # Art. 15(1)
SHORT_CHART_YEARS = 5         # Art. 15(2)
MIN_COMPLETE_YEARS = 1        # Art. 15(4)
DISPLAY_DECIMALS = 1          # Annex III §5

# --- house constants ------------------------------------------------------
RECONCILIATION_TOLERANCE_PP = 0.05   # percentage points, vs client figures
EXTREME_ANNUAL_RETURN_PCT = 75.0     # tripwire for an unadjusted / wrong series
MIN_ANCHOR_TOLERANCE_DAYS = 8        # floor, so daily and weekly series behave


# =========================================================================
# Findings
# =========================================================================

# One `Finding` class and one code enum across both KIID calculations — the
# past-performance codes are declared in `srri_engine.FindingCode` alongside
# the SRRI ones, because the user makes ONE upload and the UI renders ONE
# findings list.  The alias below keeps the references in this module reading
# as past-performance codes without introducing a second contract.
PPFindingCode = FindingCode


def _err(code: PPFindingCode, msg: str, remediation: Optional[str] = None, **detail) -> Finding:
    return Finding(code=code, severity=Severity.ERROR, message=msg,
                   remediation=remediation, detail=detail)


def _warn(code: PPFindingCode, msg: str, remediation: Optional[str] = None, **detail) -> Finding:
    return Finding(code=code, severity=Severity.WARNING, message=msg,
                   remediation=remediation, detail=detail)


def _info(code: PPFindingCode, msg: str, **detail) -> Finding:
    return Finding(code=code, severity=Severity.INFO, message=msg, detail=detail)


# =========================================================================
# Inputs
# =========================================================================

class MaterialChange(BaseModel):
    """Art. 17 — a material change to objectives and investment policy.

    Performance before the change is still shown; the earlier period carries a
    warning that it was achieved under circumstances that no longer apply.
    """
    model_config = ConfigDict(frozen=True)

    effective_date: date
    description: str = Field(min_length=3)

    @field_validator("description")
    @classmethod
    def _substantive(cls, v: str) -> str:
        if v.strip().lower() in {"change", "n/a", "na", "tbc", "-"}:
            raise ValueError(
                "Art. 17(2) requires a clear warning; give the actual change."
            )
        return v.strip()


class SimulationBasis(str, Enum):
    """Art. 19(1) — the only permitted grounds for simulated performance."""
    OTHER_SHARE_CLASS = "other_share_class"        # Art. 19(1)(a)
    MASTER_UCITS = "master_ucits"                  # Art. 19(1)(b)


class SimulatedSegment(BaseModel):
    """Declares that the series is simulated up to and including a date.

    The engine does not create simulated data.  This records that the preparer
    supplied some, so that Art. 19(2)'s prominent disclosure can be emitted and
    the affected bars flagged.
    """
    model_config = ConfigDict(frozen=True)

    through_date: date
    basis: SimulationBasis
    description: str = Field(min_length=3)


class ClientYearFigure(BaseModel):
    """An administrator-supplied yearly figure, for reconciliation only."""
    model_config = ConfigDict(frozen=True)

    year: int
    fund_return_pct: float
    benchmark_return_pct: Optional[float] = None
    source: Optional[str] = None


# =========================================================================
# Outputs
# =========================================================================

class BlankReason(str, Enum):
    FUND_DID_NOT_EXIST = "fund_did_not_exist"
    NO_DATA = "no_data"
    INCOMPLETE_YEAR = "incomplete_year"


class YearBar(BaseModel):
    """One bar on the chart.  A blank bar is still a bar — Art. 15(3) requires
    the year to appear with its date and nothing else."""
    model_config = ConfigDict(frozen=True)

    year: int

    # displayed value, rounded to Annex III §5
    fund_return_pct: Optional[float] = None
    # unrounded, kept so rounding is never applied twice
    fund_return_exact: Optional[float] = None

    benchmark_return_pct: Optional[float] = None
    benchmark_return_exact: Optional[float] = None

    start_date: Optional[date] = None
    end_date: Optional[date] = None
    start_nav: Optional[float] = None
    end_nav: Optional[float] = None
    start_anchor_lag_days: Optional[int] = None
    end_anchor_lag_days: Optional[int] = None

    is_blank: bool = False
    blank_reason: Optional[BlankReason] = None

    is_simulated: bool = False
    prior_to_material_change: bool = False

    # reconciliation against client-supplied figures
    client_return_pct: Optional[float] = None
    client_difference_pp: Optional[float] = None
    client_matches: Optional[bool] = None

    @property
    def label(self) -> str:
        """Annex III §4 — every bar labelled with the return it achieved."""
        if self.fund_return_pct is None:
            return ""
        return f"{self.fund_return_pct:.{DISPLAY_DECIMALS}f}%"


class PPStatus(str, Enum):
    OK = "ok"
    OK_WITH_WARNINGS = "ok_with_warnings"
    INSUFFICIENT_DATA = "insufficient_data"   # Art. 15(4) — statement, no chart
    BLOCKED = "blocked"


class PPAuditInfo(BaseModel):
    """Everything needed to answer "why did the 2026 KIID show 12.4 % for 2023?"
    from stored data alone, months later."""
    model_config = ConfigDict(frozen=True)

    engine_name: str = ENGINE_NAME
    engine_version: str = ENGINE_VERSION
    methodology_ref: str = METHODOLOGY_REF
    calculated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

    input_sha256: str
    input_filename: Optional[str] = None
    input_bytes: int = 0
    input_kind: str = "bytes"
    date_format_resolved: DateFormat = DateFormat.DMY

    benchmark_sha256: Optional[str] = None
    benchmark_filename: Optional[str] = None

    reference_date: date
    last_complete_year: int
    chart_years: list[int]
    chart_window_length: int
    anchor_tolerance_days: int
    currency: str
    fund_inception_date: Optional[date] = None
    benchmark_name: Optional[str] = None

    # assumptions carried, per house policy, on every single run
    assumes_net_of_ongoing_charges: bool = True
    assumes_distribution_adjusted: bool = True
    charge_adjustment_applied: bool = False

    @property
    def fingerprint(self) -> str:
        return f"{self.engine_version}/{self.input_sha256[:12]}"


class Disclosures(BaseModel):
    """The statements Art. 15(5) requires to sit alongside the chart.

    Generated as text so the document layer has one place to take them from,
    and so a missing one is visible rather than silently absent.
    """
    model_config = ConfigDict(frozen=True)

    future_performance_warning: str
    charges_statement: Optional[str] = None      # Art. 15(5)(b), conditional
    inception_statement: Optional[str] = None    # Art. 15(5)(c)
    currency_statement: str                      # Art. 15(5)(d)
    simulation_statement: Optional[str] = None   # Art. 19(2)
    material_change_statement: Optional[str] = None  # Art. 17(2)
    insufficient_data_statement: Optional[str] = None  # Art. 15(4)

    def as_list(self) -> list[str]:
        ordered = [
            self.insufficient_data_statement,
            self.future_performance_warning,
            self.charges_statement,
            self.inception_statement,
            self.currency_statement,
            self.simulation_statement,
            self.material_change_statement,
        ]
        return [s for s in ordered if s]


class PastPerformanceResult(BaseModel):
    model_config = ConfigDict(frozen=True)

    status: PPStatus
    bars: list[YearBar] = Field(default_factory=list)
    findings: list[Finding] = Field(default_factory=list)
    disclosures: Optional[Disclosures] = None
    audit: Optional[PPAuditInfo] = None

    # ---- finding views -------------------------------------------------
    @property
    def errors(self) -> list[Finding]:
        return [f for f in self.findings if f.severity == Severity.ERROR]

    @property
    def warnings(self) -> list[Finding]:
        return [f for f in self.findings if f.severity == Severity.WARNING]

    @property
    def infos(self) -> list[Finding]:
        return [f for f in self.findings if f.severity == Severity.INFO]

    @property
    def is_blocked(self) -> bool:
        return self.status == PPStatus.BLOCKED

    @property
    def requires_acknowledgement(self) -> bool:
        return bool(self.warnings)

    # ---- convenience ---------------------------------------------------
    @property
    def populated_bars(self) -> list[YearBar]:
        return [b for b in self.bars if not b.is_blank]

    @property
    def has_benchmark(self) -> bool:
        return any(b.benchmark_return_pct is not None for b in self.bars)

    def to_frame(self) -> pd.DataFrame:
        return pd.DataFrame([{
            "Year": b.year,
            "Fund return %": b.fund_return_pct,
            "Benchmark return %": b.benchmark_return_pct,
            "Start date": b.start_date,
            "Start NAV": b.start_nav,
            "End date": b.end_date,
            "End NAV": b.end_nav,
            "Blank": b.is_blank,
            "Blank reason": b.blank_reason.value if b.blank_reason else None,
            "Simulated": b.is_simulated,
            "Pre material change": b.prior_to_material_change,
            "Client %": b.client_return_pct,
            "Difference pp": b.client_difference_pp,
        } for b in self.bars])

    def summary(self) -> str:
        if self.status == PPStatus.BLOCKED:
            return "BLOCKED — " + "; ".join(f.message for f in self.errors)
        if self.status == PPStatus.INSUFFICIENT_DATA:
            return ("INSUFFICIENT DATA (Art. 15(4)) — no complete calendar year "
                    "of performance; a statement replaces the chart.")
        lines = [f"Past performance, {self.audit.currency}, "
                 f"{self.audit.chart_years[0]}–{self.audit.chart_years[-1]} "
                 f"({len(self.populated_bars)} of {len(self.bars)} years populated)"]
        for b in self.bars:
            if b.is_blank:
                lines.append(f"  {b.year}    (blank — {b.blank_reason.value})")
            else:
                bm = "" if b.benchmark_return_pct is None else \
                     f"   benchmark {b.benchmark_return_pct:+.1f}%"
                flag = "  *" if b.prior_to_material_change or b.is_simulated else ""
                lines.append(f"  {b.year}  {b.fund_return_pct:+7.1f}%{bm}{flag}")
        return "\n".join(lines)


# =========================================================================
# Core calculation
# =========================================================================

def _infer_anchor_tolerance(index: pd.DatetimeIndex) -> int:
    """Tolerance for how far before 31 December the year anchor may sit.

    Derived from the series' own cadence rather than fixed, because "8 days
    late" means something very different on a monthly file than a daily one.
    A monthly series legitimately anchors on 31 Dec; a weekly one may sit up
    to six days short.  The floor of 8 days keeps daily and weekly files from
    tripping on a normal year-end holiday run.
    """
    if len(index) < 3:
        return MIN_ANCHOR_TOLERANCE_DAYS
    spacing = np.diff(index.values).astype("timedelta64[D]").astype(int)
    median = int(np.median(spacing)) if len(spacing) else 1
    return max(MIN_ANCHOR_TOLERANCE_DAYS, median + 1)


def _year_end_anchor(prices: pd.Series, year: int) -> Optional[tuple[date, float, int]]:
    """Last observation on or before 31 December of `year`, with its lag.

    Returns (date, price, days_before_31_dec) or None where the series does not
    reach that year at all.
    """
    target = pd.Timestamp(year=year, month=12, day=31)
    window = prices.loc[:target]
    if window.empty:
        return None
    ts = window.index[-1]
    if ts.year != year:
        # The nearest earlier point belongs to a previous year — this year has
        # no data of its own.  It is a missing year, not a stale anchor.
        return None
    return ts.date(), float(window.iloc[-1]), int((target - ts).days)


def _round_display(value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    return round(value, DISPLAY_DECIMALS)


def _calendar_year_return(prices: pd.Series, year: int
                          ) -> tuple[Optional[float], Optional[tuple], Optional[tuple]]:
    """Return for calendar year `year`, plus both anchors.

    Art. 16 — based on NAV, income assumed reinvested.  With a distribution-
    adjusted series that is simply end / start - 1; no charge adjustment is
    applied (see house convention 1 in the module docstring).
    """
    start = _year_end_anchor(prices, year - 1)
    end = _year_end_anchor(prices, year)
    if start is None or end is None:
        return None, start, end
    if start[1] <= 0:
        return None, start, end
    return (end[1] / start[1] - 1.0) * 100.0, start, end


def calculate(price_series: PriceSeries,
              *,
              reference_date: Optional[date] = None,
              currency: str = "",
              fund_inception_date: Optional[date] = None,
              has_entry_or_exit_charges: bool = True,
              benchmark_series: Optional[PriceSeries] = None,
              benchmark_name: Optional[str] = None,
              client_figures: Optional[list[ClientYearFigure]] = None,
              material_changes: Optional[list[MaterialChange]] = None,
              simulated: Optional[SimulatedSegment] = None,
              anchor_tolerance_days: Optional[int] = None,
              extra_findings: Optional[list[Finding]] = None,
              ) -> PastPerformanceResult:
    """Compute the past-performance bar chart figures.

    `reference_date` is the KIID's publication reference date; everything from
    1 January of that year onward is excluded by Art. 15(6).  It defaults to
    today, which is right for an ad-hoc run and wrong for reproducing a past
    document — pass it explicitly whenever the answer must be reproducible.
    """
    # Findings raised upstream — by the pre-parse validation step — lead the
    # list. A warn-only currency mismatch that only reaches a console is a
    # warning nobody sees; carried here it lands in the audit sheet next to
    # the figures it calls into question.
    findings: list[Finding] = list(extra_findings or []) + list(price_series.findings)
    prices = price_series.prices
    ref = reference_date or date.today()

    # --- house assumptions, recorded on every run -----------------------
    findings.append(_info(
        PPFindingCode.ASSUMED_NET_OF_ONGOING_CHARGES,
        "Figures are taken from NAV and are therefore net of ongoing charges "
        "and gross of entry and exit charges. No charge adjustment applied "
        "(Art. 16).",
        charge_adjustment_applied=False))
    findings.append(_info(
        PPFindingCode.ASSUMED_DISTRIBUTION_ADJUSTED,
        "The NAV series is assumed to be accumulating or already distribution-"
        "adjusted, per house policy. The engine does not fetch or infer "
        "dividends (Art. 16 requires income to be treated as reinvested).",
        verified_by_engine=False))

    if prices.empty:
        findings.append(_err(
            PPFindingCode.NO_COMPLETE_CALENDAR_YEAR,
            "The NAV series is empty; no past performance can be produced."))
        return PastPerformanceResult(status=PPStatus.BLOCKED, findings=findings)

    series_start, series_end = prices.index[0].date(), prices.index[-1].date()

    if series_start.year > ref.year:
        findings.append(_err(
            PPFindingCode.REFERENCE_DATE_BEFORE_SERIES,
            f"The reference date {ref:%d %b %Y} precedes the start of the NAV "
            f"series ({series_start:%d %b %Y}).",
            remediation="Check the reference date and the file are the same vintage.",
            reference_date=str(ref), series_start=str(series_start)))
        return PastPerformanceResult(status=PPStatus.BLOCKED, findings=findings)

    tolerance = anchor_tolerance_days or _infer_anchor_tolerance(prices.index)
    findings.append(_info(
        PPFindingCode.ANCHOR_TOLERANCE_SET,
        f"Year-end anchors accepted up to {tolerance} days before 31 December, "
        f"derived from the series cadence.",
        anchor_tolerance_days=tolerance))

    # --- Art. 15(6): hard cut-off at 31 December of the prior year -------
    last_complete_year = ref.year - 1
    findings.append(_info(
        PPFindingCode.CURRENT_YEAR_EXCLUDED,
        f"No part of {ref.year} appears: Art. 15(6) permits no record of past "
        f"performance for any part of the current calendar year. The chart "
        f"ends at {last_complete_year}.",
        last_complete_year=last_complete_year))

    # --- Art. 23(3): the annual refresh deadline ------------------------
    # Fires only where the file genuinely fails to reach the last complete
    # year end — a weekly series closing on 26 December has not missed it.
    year_end = date(last_complete_year, 12, 31)
    if series_end < year_end - pd.Timedelta(days=tolerance):
        deadline = pd.Timestamp(year_end).date() + pd.offsets.BDay(35)
        overdue = ref > deadline.date()
        findings.append(_warn(
            PPFindingCode.STALE_PUBLICATION_WINDOW,
            f"The NAV series ends {series_end:%d %b %Y} and does not reach the "
            f"{last_complete_year} year end. Art. 23(3) requires the revised "
            f"past-performance presentation no later than 35 business days "
            f"after 31 December — {deadline.date():%d %b %Y}"
            + (f", which has passed." if overdue else "."),
            remediation="Obtain a NAV file that reaches 31 December.",
            series_end=str(series_end), deadline=str(deadline.date()),
            overdue=overdue))

    # --- which years could exist at all ---------------------------------
    inception_year = fund_inception_date.year if fund_inception_date else None
    first_possible_year = max(series_start.year + 1,
                              (inception_year + 1) if inception_year else 0)

    computed: dict[int, tuple[Optional[float], Optional[tuple], Optional[tuple]]] = {}
    for year in range(series_start.year, last_complete_year + 1):
        computed[year] = _calendar_year_return(prices, year)

    complete_years = sorted(y for y, (r, _s, _e) in computed.items() if r is not None)

    # --- Art. 15(4): nothing to show ------------------------------------
    if len(complete_years) < MIN_COMPLETE_YEARS:
        findings.append(_warn(
            PPFindingCode.NO_COMPLETE_CALENDAR_YEAR,
            "The fund has no performance data for one complete calendar year. "
            "Art. 15(4) requires a statement of insufficient data in place of "
            "the bar chart.",
            remediation="Publish the Art. 15(4) statement; do not show a chart.",
            series_start=str(series_start), series_end=str(series_end)))
        disc = _build_disclosures(
            currency=currency, fund_inception_date=fund_inception_date,
            has_entry_or_exit_charges=has_entry_or_exit_charges,
            benchmark_name=benchmark_name, simulated=simulated,
            material_changes=[], insufficient=True)
        audit = PPAuditInfo(
            input_sha256=price_series.sha256,
            input_filename=price_series.filename,
            input_bytes=price_series.n_bytes,
            input_kind=price_series.input_kind,
            date_format_resolved=price_series.date_format_resolved,
            reference_date=ref, last_complete_year=last_complete_year,
            chart_years=[], chart_window_length=0,
            anchor_tolerance_days=tolerance, currency=currency,
            fund_inception_date=fund_inception_date,
            benchmark_name=benchmark_name)
        return PastPerformanceResult(status=PPStatus.INSUFFICIENT_DATA,
                                     findings=findings, disclosures=disc,
                                     audit=audit)

    # --- Art. 15(1) / 15(2): window length ------------------------------
    window_length = (MAX_CHART_YEARS if len(complete_years) >= SHORT_CHART_YEARS
                     else SHORT_CHART_YEARS)
    if window_length == SHORT_CHART_YEARS:
        findings.append(_warn(
            PPFindingCode.FEWER_THAN_FIVE_YEARS,
            f"The fund has {len(complete_years)} complete calendar "
            f"{'year' if len(complete_years) == 1 else 'years'} of performance. "
            f"Art. 15(2) requires the presentation to cover the last 5 years "
            f"only, not 10; the remaining years appear blank.",
            complete_years=len(complete_years)))
    chart_years = list(range(last_complete_year - window_length + 1,
                             last_complete_year + 1))
    findings.append(_info(
        PPFindingCode.CHART_WINDOW_SELECTED,
        f"Chart covers {chart_years[0]}–{chart_years[-1]} "
        f"({window_length} years, Art. 15({'2' if window_length == 5 else '1'})).",
        window_length=window_length, chart_years=chart_years))

    # --- benchmark ------------------------------------------------------
    bench_prices: Optional[pd.Series] = None
    if benchmark_series is not None:
        findings.extend(benchmark_series.findings)
        if benchmark_series.prices.empty:
            findings.append(_err(
                PPFindingCode.BENCHMARK_UNUSABLE,
                "A benchmark series was supplied but contains no usable prices.",
                remediation="Art. 18(1) requires a benchmark bar alongside each "
                            "fund bar where the objectives reference one."))
            return PastPerformanceResult(status=PPStatus.BLOCKED, findings=findings)
        bench_prices = benchmark_series.prices

    # --- material changes and simulation --------------------------------
    changes = sorted(material_changes or [], key=lambda c: c.effective_date)
    in_window = [c for c in changes
                 if chart_years[0] <= c.effective_date.year <= chart_years[-1]]
    for c in in_window:
        findings.append(_warn(
            PPFindingCode.MATERIAL_CHANGE_IN_WINDOW,
            f"Material change on {c.effective_date:%d %b %Y}: {c.description}. "
            f"Art. 17 requires earlier performance to be shown and the earlier "
            f"period labelled with a warning that those circumstances no longer "
            f"apply.",
            effective_date=str(c.effective_date), description=c.description))
    earliest_change = in_window[0].effective_date if in_window else None

    if simulated is not None:
        findings.append(_warn(
            PPFindingCode.SIMULATED_PERFORMANCE_SHOWN,
            f"Performance through {simulated.through_date:%d %b %Y} is simulated "
            f"({simulated.basis.value}). Art. 19(2) requires prominent "
            f"disclosure of this on the bar chart.",
            through_date=str(simulated.through_date), basis=simulated.basis.value))

    client_map = {c.year: c for c in (client_figures or [])}

    # --- build the bars -------------------------------------------------
    bars: list[YearBar] = []
    for year in chart_years:
        ret, start, end = computed.get(year, (None, None, None))

        if ret is None:
            if inception_year is not None and year < inception_year:
                reason = BlankReason.FUND_DID_NOT_EXIST
            elif year < first_possible_year:
                reason = BlankReason.INCOMPLETE_YEAR
            else:
                reason = BlankReason.NO_DATA

            if reason is BlankReason.INCOMPLETE_YEAR:
                findings.append(_warn(
                    PPFindingCode.PARTIAL_YEAR_EXCLUDED,
                    f"{year} is not a complete calendar year for this series "
                    f"(data begins {series_start:%d %b %Y}). It is shown blank; "
                    f"a part-year return is never computed.",
                    year=year))
            elif reason is BlankReason.NO_DATA:
                findings.append(_warn(
                    PPFindingCode.MISSING_YEAR,
                    f"{year} has no usable NAV data and is shown blank with the "
                    f"date only (Art. 15(3)).",
                    year=year))
            bars.append(YearBar(year=year, is_blank=True, blank_reason=reason))
            continue

        # anchor staleness
        for lag, which in ((start[2], "opening"), (end[2], "closing")):
            if lag > tolerance:
                findings.append(_warn(
                    PPFindingCode.STALE_YEAR_ANCHOR,
                    f"The {which} anchor for {year} is {lag} days before "
                    f"31 December (tolerance {tolerance}). The bar understates "
                    f"the year by the missing period.",
                    remediation="Supply a NAV file that reaches each year end.",
                    year=year, which=which, lag_days=lag))

        if abs(ret) > EXTREME_ANNUAL_RETURN_PCT:
            findings.append(_warn(
                PPFindingCode.EXTREME_ANNUAL_RETURN,
                f"{year} returns {ret:+.1f} %, beyond the "
                f"{EXTREME_ANNUAL_RETURN_PCT:.0f} % tripwire. Check the series "
                f"is in the correct currency and is distribution-adjusted.",
                year=year, return_pct=round(ret, 4)))

        # Art. 18(2) — benchmark only for years the fund existed
        bench_ret = None
        if bench_prices is not None:
            b_ret, _bs, _be = _calendar_year_return(bench_prices, year)
            if b_ret is None:
                findings.append(_warn(
                    PPFindingCode.BENCHMARK_YEAR_MISSING,
                    f"The benchmark series has no complete {year}; the fund bar "
                    f"stands alone for that year.",
                    year=year))
            bench_ret = b_ret

        # reconciliation against client figures
        cli = client_map.get(year)
        cli_val = cli.fund_return_pct if cli else None
        diff = matches = None
        if cli_val is not None:
            diff = round(_round_display(ret) - cli_val, 4)
            matches = abs(diff) <= RECONCILIATION_TOLERANCE_PP
            if not matches:
                findings.append(_warn(
                    PPFindingCode.CLIENT_FIGURE_MISMATCH,
                    f"{year}: computed {_round_display(ret):+.1f} % against the "
                    f"supplied {cli_val:+.1f} % — a difference of {diff:+.2f} "
                    f"percentage points.",
                    remediation="Reconcile before publishing. The administrator's "
                                "figure and the NAV file disagree.",
                    year=year, computed=_round_display(ret), supplied=cli_val,
                    difference_pp=diff))

        bars.append(YearBar(
            year=year,
            fund_return_pct=_round_display(ret),
            fund_return_exact=ret,
            benchmark_return_pct=_round_display(bench_ret),
            benchmark_return_exact=bench_ret,
            start_date=start[0], end_date=end[0],
            start_nav=start[1], end_nav=end[1],
            start_anchor_lag_days=start[2], end_anchor_lag_days=end[2],
            is_simulated=bool(simulated and end[0] <= simulated.through_date),
            prior_to_material_change=bool(
                earliest_change and end[0] < earliest_change),
            client_return_pct=cli_val,
            client_difference_pp=diff,
            client_matches=matches,
        ))

    # client years that fall outside the chart
    for year in sorted(set(client_map) - set(chart_years)):
        findings.append(_warn(
            PPFindingCode.CLIENT_FIGURE_YEAR_UNMATCHED,
            f"A figure was supplied for {year}, which is outside the chart "
            f"window {chart_years[0]}–{chart_years[-1]}. It was not reconciled.",
            year=year))
    if client_map:
        reconciled = [b for b in bars if b.client_matches is not None]
        findings.append(_info(
            PPFindingCode.CLIENT_FIGURES_RECONCILED,
            f"{sum(1 for b in reconciled if b.client_matches)} of "
            f"{len(reconciled)} supplied years agree within "
            f"{RECONCILIATION_TOLERANCE_PP} pp. The engine's figures populate "
            f"the chart; the comparison is evidence, not an override.",
            tolerance_pp=RECONCILIATION_TOLERANCE_PP))

    disc = _build_disclosures(
        currency=currency, fund_inception_date=fund_inception_date,
        has_entry_or_exit_charges=has_entry_or_exit_charges,
        benchmark_name=benchmark_name, simulated=simulated,
        material_changes=in_window, insufficient=False)

    audit = PPAuditInfo(
        input_sha256=price_series.sha256,
        input_filename=price_series.filename,
        input_bytes=price_series.n_bytes,
        input_kind=price_series.input_kind,
        date_format_resolved=price_series.date_format_resolved,
        benchmark_sha256=benchmark_series.sha256 if benchmark_series else None,
        benchmark_filename=benchmark_series.filename if benchmark_series else None,
        reference_date=ref, last_complete_year=last_complete_year,
        chart_years=chart_years, chart_window_length=window_length,
        anchor_tolerance_days=tolerance, currency=currency,
        fund_inception_date=fund_inception_date, benchmark_name=benchmark_name)

    status = (PPStatus.OK_WITH_WARNINGS
              if any(f.severity == Severity.WARNING for f in findings)
              else PPStatus.OK)
    return PastPerformanceResult(status=status, bars=bars, findings=findings,
                                 disclosures=disc, audit=audit)


def _build_disclosures(*, currency: str,
                       fund_inception_date: Optional[date],
                       has_entry_or_exit_charges: bool,
                       benchmark_name: Optional[str],
                       simulated: Optional[SimulatedSegment],
                       material_changes: list[MaterialChange],
                       insufficient: bool) -> Disclosures:
    """Art. 15(5), plus Art. 17(2) and Art. 19(2) where they apply."""
    charges = None
    if has_entry_or_exit_charges:
        # Art. 15(5)(b), disapplied for funds with no entry or exit charges.
        charges = ("Past performance is calculated after ongoing charges have "
                   "been taken from the fund. It does not take account of any "
                   "entry or exit charges you may pay.")

    inception = None
    if fund_inception_date:
        inception = f"The fund came into existence in {fund_inception_date.year}."

    currency_stmt = (f"Past performance is calculated in {currency}."
                     if currency else
                     "Past performance is calculated in the fund's base currency. "
                     "[CURRENCY NOT SUPPLIED — Art. 15(5)(d) requires it.]")

    sim = None
    if simulated:
        basis = ("the performance of another share class"
                 if simulated.basis is SimulationBasis.OTHER_SHARE_CLASS
                 else "the performance of the master UCITS")
        sim = (f"Performance up to {simulated.through_date:%d %B %Y} is "
               f"simulated, based on {basis}. {simulated.description}")

    change = None
    if material_changes:
        first = material_changes[0]
        change = (f"The fund's objectives and investment policy changed on "
                  f"{first.effective_date:%d %B %Y} ({first.description}). "
                  f"Performance before that date was achieved under "
                  f"circumstances that no longer apply.")

    insufficient_stmt = None
    if insufficient:
        insufficient_stmt = ("There is insufficient data to provide a useful "
                             "indication of past performance to investors.")

    return Disclosures(
        future_performance_warning=(
            "Past performance is not a reliable indicator of future results."),
        charges_statement=charges,
        inception_statement=inception,
        currency_statement=currency_stmt,
        simulation_statement=sim,
        material_change_statement=change,
        insufficient_data_statement=insufficient_stmt,
    )


def run(source: Source,
        *,
        reference_date: Optional[date] = None,
        currency: str = "",
        fund_inception_date: Optional[date] = None,
        has_entry_or_exit_charges: bool = True,
        date_format: DateFormat = DateFormat.DMY,
        filename: Optional[str] = None,
        sheet: Optional[Union[str, int]] = None,
        skiprows: int = 0,
        date_column: Optional[str] = None,
        price_column: Optional[str] = None,
        benchmark_source: Optional[Source] = None,
        benchmark_name: Optional[str] = None,
        benchmark_filename: Optional[str] = None,
        client_figures: Optional[list[ClientYearFigure]] = None,
        material_changes: Optional[list[MaterialChange]] = None,
        simulated: Optional[SimulatedSegment] = None,
        anchor_tolerance_days: Optional[int] = None,
        ) -> PastPerformanceResult:
    """Parse, then calculate.  One call for the common case."""
    series = read_prices(source, date_format=date_format, filename=filename,
                         sheet=sheet, skiprows=skiprows,
                         date_column=date_column, price_column=price_column)
    bench = None
    if benchmark_source is not None:
        bench = read_prices(benchmark_source, date_format=date_format,
                            filename=benchmark_filename, sheet=sheet,
                            skiprows=skiprows, date_column=date_column,
                            price_column=price_column)
    return calculate(series,
                     reference_date=reference_date,
                     currency=currency,
                     fund_inception_date=fund_inception_date,
                     has_entry_or_exit_charges=has_entry_or_exit_charges,
                     benchmark_series=bench, benchmark_name=benchmark_name,
                     client_figures=client_figures,
                     material_changes=material_changes,
                     simulated=simulated,
                     anchor_tolerance_days=anchor_tolerance_days)


def run_kiid(source: Source,
             *,
             frequency: "Frequency" = None,
             reference_date: Optional[date] = None,
             currency: str = "",
             fund_inception_date: Optional[date] = None,
             has_entry_or_exit_charges: bool = True,
             date_format: DateFormat = DateFormat.DMY,
             filename: Optional[str] = None,
             sheet: Optional[Union[str, int]] = None,
             skiprows: int = 0,
             date_column: Optional[str] = None,
             price_column: Optional[str] = None,
             min_periods_override=None,
             benchmark_source: Optional[Source] = None,
             benchmark_name: Optional[str] = None,
             benchmark_filename: Optional[str] = None,
             client_figures: Optional[list[ClientYearFigure]] = None,
             material_changes: Optional[list[MaterialChange]] = None,
             simulated: Optional[SimulatedSegment] = None,
             anchor_tolerance_days: Optional[int] = None,
             extra_findings: Optional[list[Finding]] = None,
             ):
    """Both KIID calculations from ONE upload, parsed ONCE.

    The user supplies a single NAV file. The SRRI and the past-performance bar
    chart are two readings of that same series, so the file is parsed once and
    the resulting `PriceSeries` — same bytes, same SHA-256, same resolved date
    format — feeds both engines. Both results therefore carry the identical
    input hash, which is the point: one upload, one provenance record.

    Returns `(SRRIResult, PastPerformanceResult)`.

    The only separate input is an optional benchmark series, which exists
    solely because Art. 18 requires a benchmark bar where the fund's objectives
    reference one. It is not a second NAV file and is not required.

    `source` accepts raw bytes, so the upstream validation step can check the
    uploaded file and hand THE SAME BYTES straight through. That matters: if
    the validator opens the workbook and the engines then open it again, the
    audit hash no longer proves the validator and the calculation saw the same
    file.

    `extra_findings` carries that validator's own findings into both results,
    so they appear in the audit sheets rather than only in a console. A
    warn-only currency mismatch is exactly the case this exists for — the
    figures still get produced, and the warning travels with them.
    """
    import srri_engine as _srri

    freq = frequency if frequency is not None else _srri.Frequency.AUTO
    upstream = list(extra_findings or [])

    ps = read_prices(source, date_format=date_format, filename=filename,
                     sheet=sheet, skiprows=skiprows,
                     date_column=date_column, price_column=price_column)

    checks = _srri.validate(ps, freq, min_periods_override=min_periods_override)
    srri_result = _srri.calculate(ps, freq,
                                  findings=upstream + list(ps.findings) + checks,
                                  min_periods_override=min_periods_override,
                                  sheet=sheet, skiprows=skiprows)

    bench = None
    if benchmark_source is not None:
        bench = read_prices(benchmark_source, date_format=date_format,
                            filename=benchmark_filename, sheet=sheet,
                            skiprows=skiprows, date_column=date_column,
                            price_column=price_column)

    pp_result = calculate(ps,
                          reference_date=reference_date,
                          currency=currency,
                          fund_inception_date=fund_inception_date,
                          has_entry_or_exit_charges=has_entry_or_exit_charges,
                          benchmark_series=bench, benchmark_name=benchmark_name,
                          client_figures=client_figures,
                          material_changes=material_changes,
                          simulated=simulated,
                          anchor_tolerance_days=anchor_tolerance_days,
                          extra_findings=upstream)

    return srri_result, pp_result


def currency_finding(series_currency: Optional[str],
                     base_currency: Optional[str],
                     *,
                     blocking: bool = False) -> Optional[Finding]:
    """The currency guard, as a `Finding` the validator can hand to `run_kiid`.

    A NAV series quoted in anything other than the share class's base currency
    measures the fund PLUS the exchange rate. On `IE00BK5BQY34` the USD-quoted
    listing gave SRRI 6 against a published 5, and 2025 past performance of
    +36.2 % against a published +20.7 %. The bar chart makes the contamination
    far more obvious than the SRRI does.

    House setting is warn-only: the figures are still produced and the warning
    travels with them into the audit sheet. Pass `blocking=True` to make it an
    error instead, which stops both engines.

    Returns None when the two agree or either is unknown — an absent currency
    is a separate problem for the validator to raise in its own terms.
    """
    if not series_currency or not base_currency:
        return None
    if series_currency.upper() == base_currency.upper():
        return None
    return Finding(
        code=FindingCode.CURRENCY_MISMATCH,
        severity=Severity.ERROR if blocking else Severity.WARNING,
        message=(f"The NAV series is quoted in {series_currency.upper()} but the "
                 f"share class's base currency is {base_currency.upper()}. Every "
                 f"figure below embeds {series_currency.upper()}/"
                 f"{base_currency.upper()} exchange-rate movement, which is "
                 f"neither the fund's risk nor its performance."),
        remediation=("Obtain the series quoted in the base currency. Do not "
                     "publish from this one."),
        detail={"series_currency": series_currency.upper(),
                "base_currency": base_currency.upper(),
                "blocking": blocking},
    )


# ======================================================================
# Excel export — an artifact derived from a result, never the other way round
# ======================================================================

SEVERITY_HEX = {
    Severity.ERROR: "F44336",
    Severity.WARNING: "FFC107",
    Severity.INFO: "E8F0F5",
}
FUND_HEX = "1F4E78"
BENCH_HEX = "9E9E9E"


class OutputProfile(str, Enum):
    """How much of the workbook to build.

    FULL      everything the engines produce.
    KIID      the house working format, set 9 August 2026: Summary trimmed to
              the headline result, SRRI Calculations, both audit sheets, and a
              Past Performance sheet carrying the figures and the chart only.
              Dropped: Distribution, Methodology, the Summary's SRRI-
              distribution and provenance blocks, the past-performance
              methodology reference line, and the Art. 15(5) statements block.

    The statements are dropped from the WORKBOOK, not from the document. They
    remain on `result.disclosures` and Art. 15(5) still requires them to appear
    alongside the published chart — whatever renders the KIID has to take them
    from there.
    """
    FULL = "full"
    KIID = "kiid"


# Sheets the KIID profile removes from the SRRI half of the workbook.
KIID_DROP_SHEETS = ("Distribution", "Methodology")

# The Summary sheet is cut from this label to the end, which takes the SRRI
# distribution block and the provenance block below it. Matched on the label
# rather than a row number so a layout change cannot silently cut the wrong
# rows — if the label moves, the cut moves with it; if it disappears, nothing
# is cut and the sheet is left intact.
SUMMARY_CUT_FROM_LABEL = "SRRI DISTRIBUTION"


def _trim_summary_sheet(wb) -> bool:
    """Cut the Summary sheet back to the headline result. True if it cut."""
    if "Summary" not in wb.sheetnames:
        return False
    ws = wb["Summary"]
    if getattr(ws, "_charts", None):
        # delete_rows does not re-anchor charts; refuse rather than corrupt.
        log.warning("Summary sheet carries a chart; not trimming.")
        return False
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value and SUMMARY_CUT_FROM_LABEL in str(value).upper():
            first = row - 1 if row > 1 and not ws.cell(row=row - 1,
                                                       column=1).value else row
            ws.delete_rows(first, ws.max_row - first + 1)
            return True
    return False


def _add_pp_sheets(wb, result: PastPerformanceResult,
                   *,
                   include_methodology_ref: bool = True,
                   include_statements: bool = True,
                   include_audit_sheet: bool = True) -> None:
    """Append the past-performance sheets to an open workbook.

    Shared by `export_workbook` (new workbook) and `merge_into_workbook`
    (the SRRI workbook), so the two can never drift apart.

    The three flags exist because the elements they control sit ABOVE or BELOW
    the chart's data range. Removing them after the fact would mean deleting
    rows from a sheet that has a chart anchored to it, and openpyxl does not
    re-anchor charts or shift their `Reference` ranges when rows move — the
    chart would end up plotting the wrong cells. So they are never written in
    the first place, and the row arithmetic below shifts to match.
    """
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.drawing.line import LineProperties
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    body = Font(name="Arial", size=10)
    grey = PatternFill("solid", start_color="F0F0F0", end_color="F0F0F0")
    blank_fill = PatternFill("solid", start_color="FAFAFA", end_color="FAFAFA")

    def header(ws, row, labels, widths=None):
        for c, label in enumerate(labels, 1):
            cell = ws.cell(row=row, column=c, value=label)
            cell.fill, cell.font, cell.border = hdr_fill, hdr_font, border
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
        if widths:
            from openpyxl.utils import get_column_letter
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

    has_bm = result.has_benchmark
    audit = result.audit

    # ---------------------------------------------------------------- 1
    ws = wb.create_sheet("Past Performance")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Past performance"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F4E78")

    # Everything below shifts up by one when the reference line is omitted.
    off = 0 if include_methodology_ref else 1
    if include_methodology_ref:
        ws["A2"] = METHODOLOGY_REF
        ws["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")

    if result.status is PPStatus.INSUFFICIENT_DATA:
        ws.cell(row=4 - off, column=1, value="NO BAR CHART — Article 15(4)").font = \
            Font(name="Arial", size=11, bold=True, color="C00000")
        stmt = ws.cell(row=5 - off, column=1,
                       value=(result.disclosures.insufficient_data_statement
                              if result.disclosures else ""))
        stmt.font = body
        ws.column_dimensions["A"].width = 100
        if include_audit_sheet:
            _add_pp_audit_sheet(wb, result)
        return

    cols = ["Year", "Fund return %"] + (
        [f"{audit.benchmark_name or 'Benchmark'} return %"] if has_bm else []) + [
        "Opening date", "Opening NAV", "Closing date", "Closing NAV", "Note"]
    header_row = 4 - off
    header(ws, header_row, cols, widths=[10, 16] + ([22] if has_bm else []) +
           [14, 14, 14, 14, 34])

    first_data_row = header_row + 1
    for i, b in enumerate(result.bars):
        r = first_data_row + i
        ws.cell(row=r, column=1, value=b.year).font = body
        c = 2
        # A blank year is left genuinely empty so no bar is drawn, while the
        # category label still appears — Art. 15(3).
        if not b.is_blank:
            cell = ws.cell(row=r, column=c, value=b.fund_return_pct)
            cell.number_format = "0.0"
            cell.font = body
        c += 1
        if has_bm:
            if b.benchmark_return_pct is not None:
                cell = ws.cell(row=r, column=c, value=b.benchmark_return_pct)
                cell.number_format = "0.0"
                cell.font = body
            c += 1
        for value in (b.start_date, b.start_nav, b.end_date, b.end_nav):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = body
            if isinstance(value, date):
                cell.number_format = "dd/mm/yyyy"
            elif isinstance(value, float):
                cell.number_format = "0.0000"
            c += 1

        notes = []
        if b.is_blank:
            notes.append({
                BlankReason.FUND_DID_NOT_EXIST: "Fund did not exist",
                BlankReason.NO_DATA: "No data — blank, date only (Art. 15(3))",
                BlankReason.INCOMPLETE_YEAR: "Not a complete calendar year",
            }[b.blank_reason])
        if b.is_simulated:
            notes.append("Simulated (Art. 19)")
        if b.prior_to_material_change:
            notes.append("Before material change (Art. 17)")
        if b.client_matches is False:
            notes.append(f"Differs from supplied figure by "
                         f"{b.client_difference_pp:+.2f} pp")
        note = ws.cell(row=r, column=c, value="; ".join(notes))
        note.font = Font(name="Arial", size=9, italic=True,
                         color="C00000" if b.client_matches is False else "666666")
        if b.is_blank:
            for cc in range(1, c + 1):
                ws.cell(row=r, column=cc).fill = blank_fill

    last_row = first_data_row + len(result.bars) - 1

    # ---- the chart itself, to Annex III -----------------------------
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = (f"Past performance ({audit.currency})" if audit.currency
                   else "Past performance")
    chart.y_axis.title = "Annual return %"
    # Annex III §1 — linear, never logarithmic.
    chart.y_axis.scaling.logBase = None
    # Annex III §2 — the scale adapts to the span of the bars; Excel's
    # auto-scale does this and is deliberately not overridden with a fixed
    # min/max, which would compress the bars.
    chart.y_axis.majorGridlines = None
    # Annex III §3 — the X-axis sits at 0 % performance.
    chart.x_axis.crosses = "autoZero"
    chart.y_axis.crossAx = chart.x_axis.axId
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    # Annex III §4 — every bar labelled with the return achieved,
    # Annex III §5 — to one decimal place.
    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True
    chart.dLbls.numFmt = "0.0"
    chart.dLbls.showSerName = False
    chart.dLbls.showCatName = False
    chart.dLbls.showLegendKey = False
    chart.gapWidth = 60
    chart.overlap = -10 if has_bm else None
    chart.height, chart.width = 9, 20

    n_series = 2 if has_bm else 1
    data = Reference(ws, min_col=2, max_col=1 + n_series,
                     min_row=header_row, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    from openpyxl.chart.marker import DataPoint  # noqa: F401  (kept for clarity)
    for idx, hex_colour in enumerate([FUND_HEX, BENCH_HEX][:n_series]):
        s = chart.series[idx]
        s.graphicalProperties.solidFill = hex_colour
        s.graphicalProperties.line = LineProperties(noFill=True)

    if not has_bm:
        chart.legend = None

    ws.add_chart(chart, f"A{last_row + 3}")

    # ---- the Art. 15(5) statements, under the chart ------------------
    # Dropped from the workbook under the KIID profile. They are NOT dropped
    # from the disclosure: Art. 15(5) still requires them alongside the
    # published chart, and they remain on `result.disclosures` for whatever
    # renders the document.
    if not include_statements:
        if any(b.client_return_pct is not None for b in result.bars):
            _add_pp_reconciliation_sheet(wb, result, header)
        if include_audit_sheet:
            _add_pp_audit_sheet(wb, result)
        return

    stmt_row = last_row + 22
    ws.cell(row=stmt_row, column=1,
            value="Statements required alongside the chart").font = Font(
        name="Arial", size=11, bold=True, color="1F4E78")
    refs = [
        ("Art. 15(5)(a)", result.disclosures.future_performance_warning),
        ("Art. 15(5)(b)", result.disclosures.charges_statement),
        ("Art. 15(5)(c)", result.disclosures.inception_statement),
        ("Art. 15(5)(d)", result.disclosures.currency_statement),
        ("Art. 19(2)", result.disclosures.simulation_statement),
        ("Art. 17(2)", result.disclosures.material_change_statement),
    ]
    r = stmt_row + 1
    for ref, text in refs:
        if not text:
            continue
        a = ws.cell(row=r, column=1, value=ref)
        a.font = Font(name="Arial", size=9, bold=True, color="666666")
        b_ = ws.cell(row=r, column=2, value=text)
        b_.font = body
        b_.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=2, end_row=r,
                       end_column=max(6, len(cols)))
        ws.row_dimensions[r].height = 30
        r += 1

    # ---------------------------------------------------------------- 2
    if any(b.client_return_pct is not None for b in result.bars):
        _add_pp_reconciliation_sheet(wb, result, header)

    if include_audit_sheet:
        _add_pp_audit_sheet(wb, result)


def _add_pp_reconciliation_sheet(wb, result: PastPerformanceResult, header) -> None:
    """Computed figures against administrator-supplied ones."""
    from openpyxl.styles import Font, PatternFill

    body = Font(name="Arial", size=10)
    rec = wb.create_sheet("PP Reconciliation")
    rec.sheet_view.showGridLines = False
    rec["A1"] = "Computed figures against supplied figures"
    rec["A1"].font = Font(name="Arial", size=14, bold=True, color="1F4E78")
    rec["A2"] = (f"Tolerance {RECONCILIATION_TOLERANCE_PP} percentage points. "
                 f"The computed figures populate the chart; this comparison "
                 f"is evidence, not an override.")
    rec["A2"].font = Font(name="Arial", size=9, italic=True, color="666666")
    header(rec, 4, ["Year", "Computed %", "Supplied %", "Difference pp",
                    "Within tolerance"], widths=[10, 14, 14, 15, 18])
    rr = 5
    for b in result.bars:
        if b.client_return_pct is None:
            continue
        rec.cell(row=rr, column=1, value=b.year).font = body
        for col, val, fmt in ((2, b.fund_return_pct, "0.0"),
                              (3, b.client_return_pct, "0.0"),
                              (4, b.client_difference_pp, "+0.00;-0.00")):
            cell = rec.cell(row=rr, column=col, value=val)
            cell.number_format = fmt
            cell.font = body
        v = rec.cell(row=rr, column=5, value="Yes" if b.client_matches else "NO")
        v.font = Font(name="Arial", size=10, bold=not b.client_matches,
                      color="000000" if b.client_matches else "FFFFFF")
        if not b.client_matches:
            v.fill = PatternFill("solid", start_color="F44336",
                                 end_color="F44336")
        rr += 1


def _add_pp_audit_sheet(wb, result: PastPerformanceResult) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("PP Audit & Findings")
    ws.sheet_view.showGridLines = False
    body = Font(name="Arial", size=10)

    ws["A1"] = "Past performance — audit trail"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F4E78")

    a = result.audit
    rows: list[tuple[str, Any]] = [("ENGINE", "")]
    if a:
        rows += [
            ("Engine", f"{a.engine_name} v{a.engine_version}"),
            ("Methodology", a.methodology_ref),
            ("Calculated at (UTC)", a.calculated_at.strftime("%Y-%m-%d %H:%M:%S")),
            ("Fingerprint", a.fingerprint),
            ("", ""),
            ("INPUT", ""),
            ("File", a.input_filename or "(bytes)"),
            ("SHA-256", a.input_sha256),
            ("Size (bytes)", a.input_bytes),
            ("Date format resolved", a.date_format_resolved.value),
            ("Benchmark file", a.benchmark_filename or "(none)"),
            ("Benchmark SHA-256", a.benchmark_sha256 or "(none)"),
            ("", ""),
            ("PARAMETERS", ""),
            ("Reference date", str(a.reference_date)),
            ("Last complete calendar year", a.last_complete_year),
            ("Chart window", f"{a.chart_window_length} years"
                             + (f" ({a.chart_years[0]}–{a.chart_years[-1]})"
                                if a.chart_years else "")),
            ("Year-end anchor tolerance", f"{a.anchor_tolerance_days} days"),
            ("Currency", a.currency or "(NOT SUPPLIED)"),
            ("Fund inception", str(a.fund_inception_date) if a.fund_inception_date
                               else "(not supplied)"),
            ("Benchmark", a.benchmark_name or "(none)"),
            ("", ""),
            ("ASSUMPTIONS CARRIED", ""),
            ("Net of ongoing charges",
             "Yes — NAV is struck after ongoing charges accrue (Art. 16)"),
            ("Charge adjustment applied",
             "No — none is prescribed and none is applied"),
            ("Entry / exit charges",
             "Excluded — levied outside the fund, never in NAV"),
            ("Distribution-adjusted series",
             "Assumed, per house policy — NOT verified by the engine"),
        ]
    for off, (label, value) in enumerate(rows, 3):
        c1 = ws.cell(row=off, column=1, value=label)
        c2 = ws.cell(row=off, column=2, value=value)
        c1.font = (Font(name="Arial", size=10, bold=True, color="1F4E78")
                   if value == "" and label else body)
        c2.font = body
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        if value == "" and label:
            c1.fill = PatternFill("solid", start_color="E8F0F5",
                                  end_color="E8F0F5")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 86

    start = len(rows) + 5
    ws.cell(row=start, column=1, value="FINDINGS").font = Font(
        name="Arial", size=11, bold=True, color="1F4E78")
    hdr = start + 1
    for c, label in enumerate(["Severity", "Code", "Message", "Remediation"], 1):
        cell = ws.cell(row=hdr, column=c, value=label)
        cell.font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill = PatternFill("solid", start_color="1F4E78",
                                end_color="1F4E78")
    order = {Severity.ERROR: 0, Severity.WARNING: 1, Severity.INFO: 2}
    for i, f in enumerate(sorted(result.findings, key=lambda x: order[x.severity])):
        r = hdr + 1 + i
        cells = [
            ws.cell(row=r, column=1, value=f.severity.value.upper()),
            ws.cell(row=r, column=2, value=f.code.value if hasattr(f.code, "value")
                    else str(f.code)),
            ws.cell(row=r, column=3, value=f.message),
            ws.cell(row=r, column=4, value=f.remediation or ""),
        ]
        for cell in cells:
            cell.font = body
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        cells[0].fill = PatternFill("solid",
                                    start_color=SEVERITY_HEX[f.severity],
                                    end_color=SEVERITY_HEX[f.severity])
    for col, width in zip("ABCD", (12, 34, 78, 60)):
        ws.column_dimensions[col].width = width


def _profile_flags(profile: Union[OutputProfile, str]) -> dict:
    profile = OutputProfile(profile)
    if profile is OutputProfile.KIID:
        return {"include_methodology_ref": False,
                "include_statements": False,
                "include_audit_sheet": True}
    return {"include_methodology_ref": True,
            "include_statements": True,
            "include_audit_sheet": True}


def export_workbook(result: PastPerformanceResult,
                    destination: Optional[Union[str, Path, IO[bytes]]] = None,
                    *,
                    profile: Union[OutputProfile, str] = OutputProfile.FULL
                    ) -> bytes:
    """Standalone past-performance workbook.

    Sheets: Past Performance | PP Reconciliation (when supplied figures were
    given) | PP Audit & Findings.  Bytes are always returned, which is what a
    web response needs; `destination` is optional.
    """
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    _add_pp_sheets(wb, result, **_profile_flags(profile))

    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    if destination is not None:
        if hasattr(destination, "write"):
            destination.write(raw)
        else:
            Path(destination).write_bytes(raw)
    return raw


def merge_into_workbook(srri_workbook: Union[bytes, str, Path, IO[bytes]],
                        result: PastPerformanceResult,
                        destination: Optional[Union[str, Path, IO[bytes]]] = None,
                        *,
                        profile: Union[OutputProfile, str] = OutputProfile.FULL
                        ) -> bytes:
    """Append the past-performance sheets to an existing SRRI workbook.

    This is the intended output path: one upload, one workbook, the SRRI
    sheets and the past-performance sheets side by side, sharing an input hash.

    Under `profile="kiid"` the SRRI half is trimmed too — the Distribution and
    Methodology sheets are removed and the Summary sheet is cut back to the
    headline result. See `OutputProfile`.
    """
    from openpyxl import load_workbook

    src = io.BytesIO(srri_workbook) if isinstance(srri_workbook, bytes) \
        else srri_workbook
    wb = load_workbook(src)
    for name in ("Past Performance", "PP Reconciliation", "PP Audit & Findings"):
        if name in wb.sheetnames:
            wb.remove(wb[name])

    if OutputProfile(profile) is OutputProfile.KIID:
        for name in KIID_DROP_SHEETS:
            if name in wb.sheetnames:
                wb.remove(wb[name])
        _trim_summary_sheet(wb)

    _add_pp_sheets(wb, result, **_profile_flags(profile))

    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()
    if destination is not None:
        if hasattr(destination, "write"):
            destination.write(raw)
        else:
            Path(destination).write_bytes(raw)
    return raw


# ======================================================================
# CLI — a thin wrapper, not the entry point
# ======================================================================

def main(argv: Optional[list[str]] = None) -> int:
    p = argparse.ArgumentParser(
        description="UCITS KIID past-performance figures from a NAV file "
                    "(Reg. (EU) No 583/2010, Arts. 15-19, Annex III).")
    p.add_argument("input", help="NAV file — the same file the SRRI uses")
    p.add_argument("--currency", default="", help="Base currency, e.g. EUR")
    p.add_argument("--inception", default=None,
                   help="Fund inception date, YYYY-MM-DD")
    p.add_argument("--reference-date", default=None,
                   help="KIID reference date, YYYY-MM-DD (defaults to today)")
    p.add_argument("--date-format", default="dmy", choices=["dmy", "mdy", "ymd"])
    p.add_argument("--benchmark", default=None,
                   help="Optional benchmark series (Art. 18)")
    p.add_argument("--benchmark-name", default=None)
    p.add_argument("--no-entry-exit-charges", action="store_true",
                   help="Fund has no entry or exit charges — disapplies "
                        "the Art. 15(5)(b) statement")
    p.add_argument("--with-srri", action="store_true",
                   help="Also run the SRRI off the same parse and merge both "
                        "into one workbook")
    p.add_argument("--output", default=None, help="Workbook to write")
    args = p.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(message)s")
    fmt = DateFormat(args.date_format)
    ref = date.fromisoformat(args.reference_date) if args.reference_date else None
    inc = date.fromisoformat(args.inception) if args.inception else None

    if args.with_srri:
        import srri_engine as _srri
        srri_result, result = run_kiid(
            args.input, reference_date=ref, currency=args.currency,
            fund_inception_date=inc, date_format=fmt,
            has_entry_or_exit_charges=not args.no_entry_exit_charges,
            benchmark_source=args.benchmark, benchmark_name=args.benchmark_name)
        print(srri_result.summary())
        print()
    else:
        srri_result = None
        result = run(args.input, reference_date=ref, currency=args.currency,
                     fund_inception_date=inc, date_format=fmt,
                     has_entry_or_exit_charges=not args.no_entry_exit_charges,
                     benchmark_source=args.benchmark,
                     benchmark_name=args.benchmark_name)

    print(result.summary())
    for f in result.findings:
        if f.severity is not Severity.INFO:
            print(f"  {f}")

    if args.output:
        if srri_result is not None:
            import srri_engine as _srri
            merge_into_workbook(_srri.export_workbook(srri_result), result,
                                destination=args.output)
        else:
            export_workbook(result, destination=args.output)
        print(f"\nWritten: {args.output}")

    return 1 if result.is_blocked else 0


if __name__ == "__main__":       # pragma: no cover
    sys.exit(main())
